"""
S&R Extract — Unified Excel Exporter
=======================================
Produces the exact required Excel format for both Instagram and Facebook
extraction results.

Sheet layout
------------
Row 1 : Coverage label (e.g. "January – April 2025")   [merged A1:K1, light yellow]
Row 2 : Brand / page name                               [merged A2:K2, orange bg]
Row 3 : Column headers                                  [dark bg, white bold]
Row 4+: Data rows

Columns A–K:
  A  Link to Post       (clickable hyperlink)
  B  Date Posted        (MM/DD/YYYY string)
  C  No. of Comments    (integer or blank)
  D  No. of Likes / Reactions  (integer or blank)
  E  No. of Shares      (integer or blank)
  F  Total Reactions    (integer or blank; mirrors D for Instagram)
  G  POSITIVE           (integer or blank — filled from sentiment analysis)
  H  NEUTRAL            (integer or blank)
  I  NEGATIVE           (integer or blank)
  J  IRRELEVANT         (integer or blank)
  K  TOTAL              (=SUM(G{n}:J{n}) formula)

Optional Sheet 2 — "Comments":
  A  Post URL
  B  Commenter
  C  Comment Text
  D  Timestamp
  E  Sentiment          (POSITIVE / NEUTRAL / NEGATIVE / IRRELEVANT)
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_ORANGE_FILL = PatternFill("solid", fgColor="E84C1E")      # brand orange header
_YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")      # coverage label row
_DARK_FILL   = PatternFill("solid", fgColor="1F2D3D")      # column header row
_ALT_FILL    = PatternFill("solid", fgColor="F7F9FC")      # alternating row tint
_LINK_FONT   = Font(color="1155CC", underline="single", name="Calibri", size=10)
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_COVERAGE_FONT = Font(bold=True, name="Calibri", size=11, italic=True)
_BRAND_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
_BODY_FONT   = Font(name="Calibri", size=10)
_BOLD_FONT   = Font(bold=True, name="Calibri", size=10)
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=False)
_WRAP_LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
_THIN_SIDE   = Side(style="thin", color="D0D7E2")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)

COLUMNS = [
    "Link to Post",
    "Date Posted",
    "No. of Comments",
    "No. of Likes / Reactions",
    "No. of Shares",
    "Total Reactions",
    "POSITIVE",
    "NEUTRAL",
    "NEGATIVE",
    "IRRELEVANT",
    "TOTAL",
]
_NUM_COLS = len(COLUMNS)    # 11 → columns A through K
_SENTIMENT_COLS = ("G", "H", "I", "J")   # POSITIVE … IRRELEVANT
_TOTAL_COL = "K"


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _col_range(last_col_letter: str) -> str:
    return f"A1:{last_col_letter}1"


def _set_cell(ws, row: int, col: int, value: Any,
              font=None, fill=None, alignment=None, border=None,
              number_format: str | None = None) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def _auto_column_widths(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        width = max(min_width, min(max_len + 3, max_width))
        ws.column_dimensions[col_letter].width = width


def _shorten_url(url: str, max_len: int = 55) -> str:
    if len(url) <= max_len:
        return url
    return url[:max_len - 3] + "..."


def _format_date(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    return dt.strftime("%m/%d/%Y")


# ---------------------------------------------------------------------------
# Public: save posts to Sheet 1
# ---------------------------------------------------------------------------

def save_posts_excel(
    posts: list[Any],
    output_file: str,
    coverage_label: str,
    page_name: str,
    platform: str,          # "instagram" or "facebook"
) -> None:
    """
    Write (or overwrite) the main posts sheet in the given Excel file.

    ``posts`` is a list of PostData objects from either scraper:
      - Instagram: .url, .post_date_obj, .likes, .comments, .shares
      - Facebook:  .url, .post_date_obj, .reactions, .comments_count, .shares
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"
    last_col_letter = get_column_letter(_NUM_COLS)

    # ------------------------------------------------------------------
    # Row 1 — Coverage label
    # ------------------------------------------------------------------
    ws.merge_cells(f"A1:{last_col_letter}1")
    _set_cell(ws, 1, 1,
              coverage_label,
              font=_COVERAGE_FONT,
              fill=_YELLOW_FILL,
              alignment=_CENTER)
    ws.row_dimensions[1].height = 22

    # ------------------------------------------------------------------
    # Row 2 — Brand / page name (orange)
    # ------------------------------------------------------------------
    ws.merge_cells(f"A2:{last_col_letter}2")
    _set_cell(ws, 2, 1,
              page_name or platform.title(),
              font=_BRAND_FONT,
              fill=_ORANGE_FILL,
              alignment=_CENTER)
    ws.row_dimensions[2].height = 28

    # ------------------------------------------------------------------
    # Row 3 — Column headers
    # ------------------------------------------------------------------
    for col_idx, header in enumerate(COLUMNS, start=1):
        _set_cell(ws, 3, col_idx,
                  header,
                  font=_HEADER_FONT,
                  fill=_DARK_FILL,
                  alignment=_CENTER,
                  border=_THIN_BORDER)
    ws.row_dimensions[3].height = 20

    # Freeze top 3 rows and first column
    ws.freeze_panes = "B4"

    # ------------------------------------------------------------------
    # Rows 4+ — Data
    # ------------------------------------------------------------------
    is_instagram = platform.lower() == "instagram"

    for row_num, post in enumerate(posts, start=4):
        row_fill = _ALT_FILL if row_num % 2 == 0 else None

        # --- URL (col A) ---
        url = getattr(post, "url", "") or ""
        display = _shorten_url(url)
        cell_a = ws.cell(row=row_num, column=1)
        cell_a.value = display
        cell_a.hyperlink = Hyperlink(ref=f"A{row_num}", target=url, tooltip=url)
        cell_a.font = _LINK_FONT
        cell_a.alignment = _LEFT
        cell_a.border = _THIN_BORDER
        if row_fill:
            cell_a.fill = row_fill

        # --- Date (col B) ---
        date_str = _format_date(getattr(post, "post_date_obj", None))
        _set_cell(ws, row_num, 2,
                  date_str,
                  font=_BODY_FONT,
                  fill=row_fill,
                  alignment=_CENTER,
                  border=_THIN_BORDER)

        # --- Comments count (col C) ---
        comments_val: Optional[int]
        if is_instagram:
            comments_val = getattr(post, "comments", None)
        else:
            comments_val = getattr(post, "comments_count", None)
        _set_cell(ws, row_num, 3,
                  comments_val,
                  font=_BODY_FONT,
                  fill=row_fill,
                  alignment=_CENTER,
                  border=_THIN_BORDER)

        # --- Likes / Reactions (col D) ---
        likes_val: Optional[int]
        if is_instagram:
            likes_val = getattr(post, "likes", None)
        else:
            likes_val = getattr(post, "reactions", None)
        _set_cell(ws, row_num, 4,
                  likes_val,
                  font=_BODY_FONT,
                  fill=row_fill,
                  alignment=_CENTER,
                  border=_THIN_BORDER)

        # --- Shares (col E) ---
        shares_val = getattr(post, "shares", None)
        if shares_val == 0:
            shares_val = 0  # keep 0 as numeric zero, not None
        _set_cell(ws, row_num, 5,
                  shares_val,
                  font=_BODY_FONT,
                  fill=row_fill,
                  alignment=_CENTER,
                  border=_THIN_BORDER)

        # --- Total Reactions (col F) — mirrors Likes/Reactions ---
        _set_cell(ws, row_num, 6,
                  likes_val,   # IG: mirrors likes; FB: mirrors total reactions
                  font=_BODY_FONT,
                  fill=row_fill,
                  alignment=_CENTER,
                  border=_THIN_BORDER)

        # --- Sentiment cols G–J — blank initially, filled after comment collection ---
        for sent_col_idx in range(7, 11):   # G=7, H=8, I=9, J=10
            _set_cell(ws, row_num, sent_col_idx,
                      None,
                      font=_BODY_FONT,
                      fill=row_fill,
                      alignment=_CENTER,
                      border=_THIN_BORDER)

        # --- TOTAL (col K) — formula =SUM(G{n}:J{n}) ---
        formula = f"=IF(COUNTA(G{row_num}:J{row_num})>0,SUM(G{row_num}:J{row_num}),\"\")"
        _set_cell(ws, row_num, 11,
                  formula,
                  font=_BOLD_FONT,
                  fill=row_fill,
                  alignment=_CENTER,
                  border=_THIN_BORDER)

        ws.row_dimensions[row_num].height = 18

    # Column widths
    _auto_column_widths(ws)
    ws.column_dimensions["A"].width = 55   # Link column — always wide
    ws.column_dimensions["B"].width = 16   # Date
    for col_letter in ["C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 20
    for col_letter in ["G", "H", "I", "J", "K"]:
        ws.column_dimensions[col_letter].width = 14

    Path(output_file).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)


# ---------------------------------------------------------------------------
# Public: add comments sheet to existing workbook
# ---------------------------------------------------------------------------

def add_comments_sheet(
    output_file: str,
    comments: list[dict],
) -> None:
    """
    Appends or replaces a 'Comments' sheet in the existing workbook.

    ``comments`` is a list of dicts with keys:
      post_url, commenter, text, timestamp, sentiment (optional)
    """
    path = Path(output_file)
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()

    # Remove existing Comments sheet if present
    if "Comments" in wb.sheetnames:
        del wb["Comments"]

    ws = wb.create_sheet("Comments")

    # Header row
    comment_headers = ["Post URL", "Commenter", "Comment Text", "Timestamp", "Sentiment"]
    for col_idx, header in enumerate(comment_headers, start=1):
        _set_cell(ws, 1, col_idx,
                  header,
                  font=_HEADER_FONT,
                  fill=_DARK_FILL,
                  alignment=_CENTER,
                  border=_THIN_BORDER)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Data rows
    for row_num, comment in enumerate(comments, start=2):
        row_fill = _ALT_FILL if row_num % 2 == 0 else None

        post_url = str(comment.get("post_url") or "")
        commenter = str(comment.get("commenter") or "")
        text = str(comment.get("text") or "")
        timestamp = str(comment.get("timestamp") or "")
        sentiment = str(comment.get("sentiment") or "")

        # Post URL — clickable
        cell_a = ws.cell(row=row_num, column=1)
        cell_a.value = _shorten_url(post_url)
        if post_url:
            cell_a.hyperlink = Hyperlink(ref=f"A{row_num}", target=post_url, tooltip=post_url)
            cell_a.font = _LINK_FONT
        else:
            cell_a.font = _BODY_FONT
        cell_a.alignment = _LEFT
        cell_a.border = _THIN_BORDER
        if row_fill:
            cell_a.fill = row_fill

        _set_cell(ws, row_num, 2, commenter,  font=_BODY_FONT, fill=row_fill, alignment=_LEFT,   border=_THIN_BORDER)
        _set_cell(ws, row_num, 3, text,       font=_BODY_FONT, fill=row_fill, alignment=_WRAP_LEFT, border=_THIN_BORDER)
        _set_cell(ws, row_num, 4, timestamp,  font=_BODY_FONT, fill=row_fill, alignment=_CENTER, border=_THIN_BORDER)

        # Sentiment cell — colour-coded
        sent_fill = None
        if sentiment == "POSITIVE":
            sent_fill = PatternFill("solid", fgColor="C6EFCE")
        elif sentiment == "NEGATIVE":
            sent_fill = PatternFill("solid", fgColor="FFC7CE")
        elif sentiment == "NEUTRAL":
            sent_fill = PatternFill("solid", fgColor="FFEB9C")
        elif sentiment == "IRRELEVANT":
            sent_fill = PatternFill("solid", fgColor="E0E0E0")
        _set_cell(ws, row_num, 5, sentiment,
                  font=_BOLD_FONT,
                  fill=sent_fill or row_fill,
                  alignment=_CENTER,
                  border=_THIN_BORDER)

        ws.row_dimensions[row_num].height = 30

    # Widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14

    wb.save(path)


# ---------------------------------------------------------------------------
# Public: fill sentiment counts back into the Posts sheet
# ---------------------------------------------------------------------------

def update_sentiment_counts(
    output_file: str,
    post_urls: list[str],
    comments: list[dict],
) -> None:
    """
    After comment collection, count sentiments per post and write back
    the POSITIVE/NEUTRAL/NEGATIVE/IRRELEVANT counts into the Posts sheet.

    ``post_urls`` must be the same ordered list used during extraction.
    ``comments`` must have 'post_url' and 'sentiment' keys.
    """
    from collections import defaultdict, Counter

    path = Path(output_file)
    if not path.exists():
        return

    wb = load_workbook(path)
    if "Posts" not in wb.sheetnames:
        wb.save(path)
        return

    ws = wb["Posts"]

    # Build mapping: normalised post url → sentiment counter
    url_sentiment: dict[str, Counter] = defaultdict(Counter)
    for comment in comments:
        url = str(comment.get("post_url") or "").strip().rstrip("/")
        sentiment = str(comment.get("sentiment") or "NEUTRAL")
        url_sentiment[url][sentiment] += 1

    # Data starts at row 4 (row 1 = coverage, row 2 = brand, row 3 = headers)
    url_to_data_row: dict[str, int] = {}
    for row_num, post_url in enumerate(post_urls, start=4):
        key = post_url.strip().rstrip("/")
        url_to_data_row[key] = row_num

    sentiment_order = ["POSITIVE", "NEUTRAL", "NEGATIVE", "IRRELEVANT"]
    col_letters    = ["G",         "H",        "I",         "J"]

    for url, row_num in url_to_data_row.items():
        counts = url_sentiment.get(url, Counter())
        for sentiment, col_letter in zip(sentiment_order, col_letters):
            col_idx = ord(col_letter) - ord("A") + 1
            val = counts.get(sentiment, 0) or None   # keep 0 as None to stay blank if no comments
            ws.cell(row=row_num, column=col_idx).value = val

    wb.save(path)
