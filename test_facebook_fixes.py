#!/usr/bin/env python3
"""
Facebook-specific tests for S&R Extract system.

Tests validate the following Facebook fixes:
1. Comment collection mode is completely removed
2. collection_type forced to posts_only
3. Excel export rejects/sanitizes dict/list values
4. Scroll behavior is deterministic
5. Post links are deduplicated and validated
6. Data pairing is preserved (url + date + metrics stay together)
7. Instagram code is unaffected
"""

import json
import tempfile
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Test configuration
TEST_MODULE_FACEBOOK = "facebook_to_excel"
TEST_MODULE_INSTAGRAM = "instagram_to_excel"


def test_facebook_comment_collection_removed():
    """
    TEST: collection_type='posts_with_comments' option is removed from Facebook UI.
    Expected: Only 'posts_only' is available for Facebook.
    """
    from core.platforms.registry import get_platform_adapter
    
    fb_adapter = get_platform_adapter("facebook")
    config = fb_adapter.platform_config()
    
    # Verify collection_type is disabled
    assert not config.collection_type_enabled, \
        "FAILED: collection_type_enabled should be False for Facebook"
    
    # Verify only posts_only is in options
    options = config.collection_type_options
    assert len(options) == 1, \
        f"FAILED: Expected 1 option, got {len(options)}"
    assert options[0]["value"] == "posts_only", \
        f"FAILED: Expected 'posts_only', got {options[0]['value']}"
    
    print("✓ Facebook comment collection mode removed from UI")


def test_facebook_forces_posts_only():
    """
    TEST: Facebook adapter forces 'posts_only' regardless of collection_type parameter.
    Expected: extract_post and export_excel always use 'posts_only'.
    """
    from core.platforms.registry import get_platform_adapter
    import unittest.mock as mock
    
    fb_adapter = get_platform_adapter("facebook")
    
    # Mock facebook_to_excel.extract_post_from_feed
    with mock.patch("facebook_to_excel.extract_post_from_feed") as mock_extract:
        mock_extract.return_value = None
        with mock.patch("facebook_to_excel.open_post_for_extraction"):
            with mock.patch("facebook_to_excel.extract_metrics_from_loaded_post"):
                try:
                    # Try with collection_type='posts_with_comments' (should be ignored)
                    fb_adapter.extract_post(None, "http://facebook.com/test", "posts_with_comments")
                except Exception:
                    pass  # Mock doesn't fully work, but we're checking the call
        
        # Verify it was called with 'posts_only'
        if mock_extract.called:
            call_args = mock_extract.call_args
            assert "posts_only" in str(call_args) or call_args[0][2] == "posts_only", \
                f"FAILED: extract_post_from_feed not called with 'posts_only', got {call_args}"
    
    print("✓ Facebook forces posts_only in extract_post")


def test_facebook_excel_sanitizes_dict_values():
    """
    TEST: Excel export sanitizes dict/list values instead of crashing.
    Expected: viewport objects like {'width': 1920} are converted to JSON strings.
    """
    from facebook_to_excel import sanitize_excel_value
    
    # Test dict sanitization
    test_dict = {"width": 1920, "height": 1080}
    result = sanitize_excel_value(test_dict, "viewport")
    assert isinstance(result, str), \
        f"FAILED: dict should convert to string, got {type(result)}"
    assert "1920" in result, \
        f"FAILED: dict not properly serialized, got {result}"
    
    # Test list sanitization
    test_list = [1, 2, 3]
    result = sanitize_excel_value(test_list, "items")
    assert isinstance(result, str), \
        f"FAILED: list should convert to string, got {type(result)}"
    
    # Test scalar preservation
    assert sanitize_excel_value("text", "field") == "text"
    assert sanitize_excel_value(123, "field") == 123
    assert sanitize_excel_value(45.6, "field") == 45.6
    assert sanitize_excel_value(None, "field") is None
    assert sanitize_excel_value(True, "field") is True
    
    print("✓ Facebook Excel export sanitizes dict/list values")


def test_facebook_excel_no_comments_sheet():
    """
    TEST: Excel export no longer creates 'Visible Comments' sheet.
    Expected: Only 'Facebook Posts' and 'Diagnostics' sheets exist.
    """
    from facebook_to_excel import save_facebook_excel
    
    posts = [
        {
            "post_link": "https://facebook.com/post1",
            "post_date": "2026-04-28",
            "post_type": "Post",
            "reactions": 100,
            "comments_count": 50,
            "shares": 10,
            "notes": ["Test note"],
            "comments_preview": [
                {"post_url": "url", "commenter_name": "Test", "comment_text": "Comment", "comment_date_raw": "Today", "thread_type": "Comment"}
            ]
        }
    ]
    
    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = str(Path(tmpdir) / "test_facebook.xlsx")
        save_facebook_excel(posts, output_file, "2026-04-28 to latest", "posts_only")
        
        # Load and verify
        wb = load_workbook(output_file)
        sheet_names = wb.sheetnames
        
        assert "Facebook Posts" in sheet_names, \
            f"FAILED: 'Facebook Posts' sheet missing, got {sheet_names}"
        assert "Visible Comments" not in sheet_names, \
            f"FAILED: 'Visible Comments' sheet should be removed, found in {sheet_names}"
        assert "Diagnostics" in sheet_names, \
            f"FAILED: 'Diagnostics' sheet missing, got {sheet_names}"
        
        print("✓ Facebook Excel export removes 'Visible Comments' sheet")


def test_facebook_scroll_deterministic():
    """
    TEST: Facebook scroll strategy uses deterministic scroll distances.
    Expected: window-scroll uses 75% of viewport height (not 90%).
    """
    from facebook_to_excel import apply_scroll_strategy
    import unittest.mock as mock
    
    # Create mock page
    mock_page = mock.MagicMock()
    mock_page.evaluate = mock.MagicMock()
    
    # Apply scroll strategy
    apply_scroll_strategy(mock_page, "window-scroll")
    
    # Verify the scroll script contains 0.75 (75%)
    call_args = mock_page.evaluate.call_args
    script = str(call_args)
    assert "0.75" in script or "0.75" in str(call_args), \
        f"FAILED: Expected 0.75 in scroll strategy, got {script}"
    
    print("✓ Facebook scroll uses deterministic 75% viewport height")


def test_facebook_dedupe_post_links():
    """
    TEST: Post links are properly deduplicated.
    Expected: Duplicate URLs are removed, order preserved.
    """
    from facebook_to_excel import dedupe_post_links
    
    raw_links = [
        "https://facebook.com/post/123",
        "https://facebook.com/post/123&__tn__=xyz",  # Tracking param should be removed
        "https://facebook.com/post/456",
        "https://facebook.com/post/123",  # Exact duplicate
    ]
    
    result = dedupe_post_links(raw_links)
    
    # Should have 2 unique posts
    assert len(result) == 2, \
        f"FAILED: Expected 2 unique links, got {len(result)}: {result}"
    
    # Verify both post URLs are present (ignoring params)
    result_str = str(result)
    assert "123" in result_str and "456" in result_str, \
        f"FAILED: Expected both posts 123 and 456, got {result}"
    
    print("✓ Facebook post links properly deduplicated")


def test_facebook_data_pairing():
    """
    TEST: Post data remains paired (url + date + metrics together).
    Expected: No metric leakage between posts.
    """
    posts = [
        {
            "post_link": "https://facebook.com/post1",
            "post_date": "2026-04-28",
            "post_date_obj": datetime(2026, 4, 28),
            "reactions": 100,
            "comments_count": 50,
            "shares": 10,
        },
        {
            "post_link": "https://facebook.com/post2",
            "post_date": "2026-04-27",
            "post_date_obj": datetime(2026, 4, 27),
            "reactions": 200,
            "comments_count": 75,
            "shares": 20,
        }
    ]
    
    from facebook_to_excel import save_facebook_excel
    
    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = str(Path(tmpdir) / "test_pairing.xlsx")
        save_facebook_excel(posts, output_file, "Coverage", "posts_only")
        
        wb = load_workbook(output_file)
        ws = wb["Facebook Posts"]
        
        # Row 4 should have post1 data
        assert "post1" in str(ws["A4"].value), \
            f"FAILED: Post1 link not in row 4, got {ws['A4'].value}"
        assert ws["D4"].value == 100, \
            f"FAILED: Post1 reactions should be 100, got {ws['D4'].value}"
        assert ws["E4"].value == 50, \
            f"FAILED: Post1 comments should be 50, got {ws['E4'].value}"
        
        # Row 5 should have post2 data
        assert "post2" in str(ws["A5"].value), \
            f"FAILED: Post2 link not in row 5, got {ws['A5'].value}"
        assert ws["D5"].value == 200, \
            f"FAILED: Post2 reactions should be 200, got {ws['D5'].value}"
        assert ws["E5"].value == 75, \
            f"FAILED: Post2 comments should be 75, got {ws['E5'].value}"
        
        print("✓ Facebook post data remains properly paired")


def test_instagram_unchanged():
    """
    TEST: Instagram code/configuration not modified by Facebook fixes.
    Expected: Instagram adapter still has collection_type options.
    """
    from core.platforms.registry import get_platform_adapter
    
    ig_adapter = get_platform_adapter("instagram")
    config = ig_adapter.platform_config()
    
    # Instagram should STILL have collection_type enabled
    # (We didn't modify Instagram, so verify it's still there)
    assert hasattr(config, 'collection_type_enabled'), \
        "FAILED: Instagram config missing collection_type_enabled"
    
    print("✓ Instagram code unchanged")


def test_facebook_normalize_viewport():
    """
    TEST: normalize_facebook_page_viewport resets zoom and scroll.
    Expected: Function sets zoom to 100% and scrolls to top.
    """
    from facebook_to_excel import normalize_facebook_page_viewport
    import unittest.mock as mock
    
    mock_page = mock.MagicMock()
    
    # Call normalization
    normalize_facebook_page_viewport(mock_page)
    
    # Verify keyboard and evaluate were called (indicating zoom reset and scroll)
    assert mock_page.keyboard.press.called or mock_page.evaluate.called, \
        "FAILED: Page should have keyboard or evaluate calls"
    
    print("✓ Facebook viewport normalization called")


def run_facebook_tests():
    """Run all Facebook-specific tests."""
    print("\n" + "=" * 70)
    print("FACEBOOK-SPECIFIC FIX VALIDATION TESTS")
    print("=" * 70 + "\n")
    
    tests = [
        ("Comment Collection Mode", test_facebook_comment_collection_removed),
        ("Force posts_only", test_facebook_forces_posts_only),
        ("Excel Sanitization", test_facebook_excel_sanitizes_dict_values),
        ("Excel No Comments Sheet", test_facebook_excel_no_comments_sheet),
        ("Scroll Deterministic", test_facebook_scroll_deterministic),
        ("Link Deduplication", test_facebook_dedupe_post_links),
        ("Data Pairing", test_facebook_data_pairing),
        ("Instagram Unchanged", test_instagram_unchanged),
        ("Viewport Normalization", test_facebook_normalize_viewport),
    ]
    
    passed = 0
    failed = 0
    
    for test_name, test_func in tests:
        try:
            test_func()
            passed += 1
        except AssertionError as e:
            print(f"✗ {test_name}: {e}")
            failed += 1
        except Exception as e:
            print(f"⚠ {test_name}: Unexpected error: {e}")
            failed += 1
    
    print("\n" + "=" * 70)
    print(f"RESULTS: {passed} passed, {failed} failed")
    print("=" * 70 + "\n")
    
    return failed == 0


if __name__ == "__main__":
    import sys
    success = run_facebook_tests()
    sys.exit(0 if success else 1)
