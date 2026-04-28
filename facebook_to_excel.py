from __future__ import annotations

import os
import re
import subprocess
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Optional
from urllib.parse import parse_qsl, quote, urlencode, urlparse, urlunparse

from openpyxl import Workbook
from playwright.sync_api import Error as PlaywrightError


DATE_INPUT_FORMAT = "%Y-%m-%d"
VALID_FACEBOOK_HOSTS = {"facebook.com", "www.facebook.com", "m.facebook.com", "fb.com", "www.fb.com"}
INVALID_FILENAME_CHARS = set('<>:"/\\|?*')
BLOCKED_RESOURCE_TYPES = {"image", "media", "font"}
VERIFICATION_URL_TOKENS = [
    "/checkpoint/",
    "/checkpoint",
    "/login/identify",
    "/two_step_verification",
    "/authentication/",
    "/authentication",
    "/authentication?",
    "/captcha/",
    "recaptcha",
]

PLAYWRIGHT_STORAGE_STATE = os.getenv("PLAYWRIGHT_STORAGE_STATE", "").strip() or None
DEFAULT_STORAGE_STATE_FILE = Path("storage_states/facebook_auth.json")
DEFAULT_USER_DATA_DIR = Path("storage_states/facebook_user_data")
PLAYWRIGHT_HEADLESS = os.getenv("PLAYWRIGHT_HEADLESS", "true").strip().lower() not in {"0", "false", "no", "off"}
PLAYWRIGHT_AUTO_INSTALL = os.getenv("PLAYWRIGHT_AUTO_INSTALL", "true").strip().lower() not in {"0", "false", "no", "off"}
RUNNING_ON_RENDER = bool(os.getenv("RENDER") or os.getenv("RENDER_SERVICE_ID") or os.getenv("RENDER_EXTERNAL_URL"))
HAS_LOCAL_DESKTOP = os.name == "nt" or bool(os.getenv("DISPLAY") or os.getenv("WAYLAND_DISPLAY"))
PLAYWRIGHT_INTERACTIVE_BROWSER = os.getenv(
    "PLAYWRIGHT_INTERACTIVE_BROWSER",
    "true" if (HAS_LOCAL_DESKTOP and not RUNNING_ON_RENDER) else "false",
).strip().lower() in {"1", "true", "yes", "on"}
PLAYWRIGHT_BROWSER_CHANNEL = os.getenv(
    "PLAYWRIGHT_BROWSER_CHANNEL",
    "chrome" if (HAS_LOCAL_DESKTOP and not RUNNING_ON_RENDER) else "",
).strip().lower()
ACTIVE_BROWSER_ENGINE = "chromium"

# -----------------------------------------------------------------------------
# Facebook login credentials
# Prefer environment variables in production or shared environments:
#   FACEBOOK_USERNAME
#   FACEBOOK_PASSWORD
#
# PUT TEST ACCOUNT LOGIN DETAILS HERE only if you intentionally need a local
# development fallback. Keep them backend-only and never expose them to JS/HTML.
# -----------------------------------------------------------------------------
TEST_FACEBOOK_USERNAME = ""
TEST_FACEBOOK_PASSWORD = ""
FACEBOOK_USERNAME = os.getenv("FACEBOOK_USERNAME", "").strip() or TEST_FACEBOOK_USERNAME.strip()
FACEBOOK_PASSWORD = os.getenv("FACEBOOK_PASSWORD", "").strip() or TEST_FACEBOOK_PASSWORD.strip()

POST_GOTO_TIMEOUT = 35000
LOGIN_FORM_TIMEOUT = 12000
LOGIN_READY_TIMEOUT = 180000
LOGIN_POST_SUBMIT_TIMEOUT = 15000
PAGE_READY_TIMEOUT = 12000
SCROLL_WAIT_TIMEOUT = 2200
PROFILE_RETRY_MS = 800
BASE_POST_DELAY = 0.15
SLOW_SCROLL_SECONDS = 2.0
SLOW_POST_SECONDS = 4.0
COMMENT_LOAD_WAIT_MS = 900
COMMENT_EXPANSION_ROUNDS = 4

POST_LINK_SELECTOR = (
    "a[href*='/posts/'], "
    "a[href*='/videos/'], "
    "a[href*='/permalink/'], "
    "a[href*='story_fbid='], "
    "a[href*='/photo.php'], "
    "a[href*='/watch/?v=']"
)
FEED_READY_SELECTOR = "div[role='feed'], div[role='feed'] div[role='article']"
PAGE_READY_SELECTOR = f"{POST_LINK_SELECTOR}, {FEED_READY_SELECTOR}"
PAGE_SHELL_READY_SELECTOR = (
    "div[role='main'] h1, "
    "div[role='main'] div[role='tablist'], "
    "div[role='main'] a[href*='/about'], "
    "div[role='main'] a[href*='/followers'], "
    "div[role='main'] a[href*='/photos'], "
    "div[role='main'] a[href*='/videos']"
)
LOGIN_FORM_SELECTOR = "input[name='email'], input[name='pass'], form[action*='login']"

LogHook = Callable[[str, str, str], None]
ProgressHook = Callable[[int, int, int], None]


class AuthRequiredError(RuntimeError):
    def __init__(self, state: str, reason: str) -> None:
        super().__init__(reason)
        self.state = state
        self.reason = reason


@dataclass
class CommentData:
    post_url: str
    commenter_name: str
    comment_text: str
    comment_date_raw: str = ""
    thread_type: str = "Comment"


@dataclass
class PostData:
    url: str
    post_type: str
    post_date_raw: str
    post_date_obj: Optional[datetime]
    reactions: Optional[int]
    comments_count: Optional[int]
    shares: Optional[int]
    notes: str = ""
    comments_preview: list[CommentData] = field(default_factory=list)


def emit_log(log_hook: Optional[LogHook], level: str, action: str, details: str = "") -> None:
    if log_hook is None:
        return
    try:
        log_hook(level, action, details)
    except Exception:
        pass


def emit_progress(progress_hook: Optional[ProgressHook], scroll_round: int, total_rounds: int, posts_found: int) -> None:
    if progress_hook is None:
        return
    try:
        progress_hook(scroll_round, total_rounds, posts_found)
    except Exception:
        pass


def apply_context_preferences(context) -> None:
    try:
        context.add_init_script(
            """() => {
                try {
                    Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                    Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] });
                    Object.defineProperty(navigator, 'platform', { get: () => 'Win32' });
                    if (window.Notification) {
                        try {
                            Object.defineProperty(Notification, 'permission', {
                                configurable: true,
                                get: () => 'denied',
                            });
                        } catch (error) {}
                        try {
                            Notification.requestPermission = () => Promise.resolve('denied');
                        } catch (error) {}
                    }
                } catch (error) {}
            }"""
        )
    except Exception:
        pass


def uses_local_browser_window() -> bool:
    return PLAYWRIGHT_INTERACTIVE_BROWSER and HAS_LOCAL_DESKTOP and not RUNNING_ON_RENDER


def browser_runtime_mode() -> str:
    return "local_window" if uses_local_browser_window() else "headless_session"


def browser_mode_label() -> str:
    return "Opened Browser Window" if uses_local_browser_window() else "Headless Browser Session"


def browser_engine_label() -> str:
    engine = (ACTIVE_BROWSER_ENGINE or "chromium").strip().lower()
    if engine == "chrome":
        return "Google Chrome"
    if engine == "msedge":
        return "Microsoft Edge"
    return "Chromium"


def browser_mode_note() -> str:
    if uses_local_browser_window():
        return (
            f"A real {browser_engine_label()} window opens locally for Facebook login and verification "
            "using a persistent local browser profile. Keep that browser open, avoid refreshing the page, "
            "and click GO only after the target page is visible."
        )
    return "This environment cannot open a local Chromium window. Manual Facebook login requires a local run with PLAYWRIGHT_INTERACTIVE_BROWSER=true."


def preview_input_supported() -> bool:
    return False


def get_storage_state_path(require_exists: bool = False) -> Optional[Path]:
    if PLAYWRIGHT_STORAGE_STATE:
        path = Path(PLAYWRIGHT_STORAGE_STATE)
    elif uses_local_browser_window():
        path = DEFAULT_STORAGE_STATE_FILE
    else:
        return None
    if require_exists and not path.exists():
        return None
    return path


def has_saved_storage_state() -> bool:
    return get_storage_state_path(require_exists=True) is not None


def storage_state_label() -> str:
    path = get_storage_state_path()
    return str(path) if path is not None else ""


def get_user_data_dir() -> Optional[Path]:
    if not uses_local_browser_window():
        return None
    return DEFAULT_USER_DATA_DIR


def has_login_credentials() -> bool:
    return bool(FACEBOOK_USERNAME and FACEBOOK_PASSWORD)


def load_or_create_context(browser):
    # Use consistent viewport for stable positioning and scrolling
    context_options = {
        "viewport": {"width": 1365, "height": 900},
        "locale": "en-US",
        "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    }
    storage_path = get_storage_state_path(require_exists=True)
    if storage_path is not None:
        context_options["storage_state"] = str(storage_path)
    context = browser.new_context(**context_options)
    apply_context_preferences(context)
    return context, storage_path


def wait_for_selector(page, selector: str, timeout_ms: int) -> bool:
    try:
        page.locator(selector).first.wait_for(timeout=timeout_ms)
        return True
    except Exception:
        return False


def normalize_excel_filename(raw_value: str) -> Optional[str]:
    value = raw_value.strip()
    if not value:
        return None
    if not value.lower().endswith(".xlsx"):
        value = f"{value}.xlsx"
    if any(char in INVALID_FILENAME_CHARS for char in value):
        return None
    return value


def normalize_facebook_target_url(raw_value: str) -> Optional[str]:
    value = raw_value.strip()
    if not value:
        return None
    if not re.match(r"^https?://", value, re.IGNORECASE):
        value = f"https://{value}"
    try:
        parsed = urlparse(value)
    except Exception:
        return None

    hostname = (parsed.hostname or "").lower()
    if parsed.scheme not in {"http", "https"} or hostname not in VALID_FACEBOOK_HOSTS:
        return None

    if not parsed.path or parsed.path == "/":
        return None

    normalized_path = parsed.path.rstrip("/") or "/"
    return urlunparse(("https", "www.facebook.com", normalized_path, "", parsed.query, ""))


def format_date_coverage(start_date: datetime, end_date: Optional[datetime]) -> str:
    if end_date is None:
        return f"{start_date.strftime(DATE_INPUT_FORMAT)} to latest visible content"
    return f"{start_date.strftime(DATE_INPUT_FORMAT)} to {end_date.strftime(DATE_INPUT_FORMAT)}"


def parse_count(value: Optional[str]) -> Optional[int]:
    if not value:
        return None
    text = value.strip().upper().replace(",", "")
    match = re.match(r"^(\d+(?:\.\d+)?)([KMB]?)$", text)
    if match:
        number = float(match.group(1))
        suffix = match.group(2)
        if suffix == "K":
            number *= 1_000
        elif suffix == "M":
            number *= 1_000_000
        elif suffix == "B":
            number *= 1_000_000_000
        return int(number)

    digits = re.sub(r"[^\d]", "", text)
    return int(digits) if digits else None


def get_target_page_slug(target_url: str) -> str:
    try:
        parsed = urlparse(target_url)
    except Exception:
        return ""
    path = (parsed.path or "").strip("/")
    if not path:
        return ""
    return path.split("/")[0].strip().lower()


def normalize_facebook_post_url(raw_value: str, target_url: str = "") -> Optional[str]:
    if not raw_value:
        return None

    try:
        parsed = urlparse(raw_value)
    except Exception:
        return None

    host = (parsed.hostname or "").lower()
    if host and host not in VALID_FACEBOOK_HOSTS:
        return None

    path = (parsed.path or "").rstrip("/")
    lower_path = path.lower()
    if not any(marker in lower_path for marker in ("/posts/", "/videos/", "/permalink/", "/watch/", "/photo.php")) and "story_fbid=" not in (parsed.query or "").lower():
        return None

    target_slug = get_target_page_slug(target_url)
    if target_slug and path.strip("/"):
        first_segment = path.strip("/").split("/")[0].strip().lower()
        slug_style = any(marker in lower_path for marker in ("/posts/", "/videos/", "/photos/", "/reels/", "/permalink/"))
        if slug_style and first_segment and first_segment not in {target_slug, "photo.php", "watch", "permalink.php"}:
            # Allow broader matches for public content; log-level filtering happens later.
            pass

    keep_query_keys = {"story_fbid", "id", "fbid", "set", "type", "theater", "v"}
    query_pairs = []
    for key, value in parse_qsl(parsed.query, keep_blank_values=False):
        key_lower = key.lower()
        if key_lower in {"comment_id", "reply_comment_id", "notif_id", "notif_t", "ref", "__cft__", "__tn__"}:
            continue
        if key_lower in keep_query_keys:
            query_pairs.append((key, value))

    query = urlencode(query_pairs, doseq=True)
    normalized = parsed._replace(params="", query=query, fragment="")
    return urlunparse(normalized)


def dedupe_post_links(raw_links: list[str], target_url: str = "") -> list[str]:
    seen: dict[str, bool] = {}
    for raw_value in raw_links:
        normalized = normalize_facebook_post_url(raw_value, target_url=target_url)
        if not normalized or normalized in seen:
            continue
        seen[normalized] = True
    return list(seen.keys())


def parse_facebook_datetime(raw_value: str) -> Optional[datetime]:
    if not raw_value:
        return None

    text = raw_value.strip()
    normalized = re.sub(r"\s+", " ", text.lower()).strip()
    now = datetime.now()

    short_relative_match = re.fullmatch(r"(\d+)\s*([smhdw])", normalized)
    if short_relative_match:
        amount = int(short_relative_match.group(1))
        unit = short_relative_match.group(2)
        if unit == "s":
            return now - timedelta(seconds=amount)
        if unit == "m":
            return now - timedelta(minutes=amount)
        if unit == "h":
            return now - timedelta(hours=amount)
        if unit == "d":
            return now - timedelta(days=amount)
        if unit == "w":
            return now - timedelta(weeks=amount)

    relative_match = re.search(r"(\d+)\s+(second|minute|hour|day|week|month|year)s?\s+ago", normalized)
    if relative_match:
        amount = int(relative_match.group(1))
        unit = relative_match.group(2)
        if unit == "second":
            return now - timedelta(seconds=amount)
        if unit == "minute":
            return now - timedelta(minutes=amount)
        if unit == "hour":
            return now - timedelta(hours=amount)
        if unit == "day":
            return now - timedelta(days=amount)
        if unit == "week":
            return now - timedelta(weeks=amount)
        if unit == "month":
            return now - timedelta(days=amount * 30)
        if unit == "year":
            return now - timedelta(days=amount * 365)

    if normalized.startswith("today"):
        return now

    yesterday_match = re.search(r"yesterday(?: at (.+))?$", normalized)
    if yesterday_match:
        base = now - timedelta(days=1)
        time_part = (yesterday_match.group(1) or "").strip()
        if time_part:
            for time_fmt in ("%I:%M %p", "%I %p"):
                try:
                    time_value = datetime.strptime(time_part.upper(), time_fmt)
                    return base.replace(hour=time_value.hour, minute=time_value.minute, second=0, microsecond=0)
                except ValueError:
                    continue
        return base

    formats = [
        "%A, %B %d, %Y at %I:%M %p",
        "%B %d, %Y at %I:%M %p",
        "%B %d at %I:%M %p",
        "%b %d, %Y at %I:%M %p",
        "%B %d, %Y",
        "%b %d, %Y",
        "%B %d",
        "%b %d",
        "%Y-%m-%d",
    ]
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            if parsed.year == 1900:
                parsed = parsed.replace(year=now.year)
            return parsed
        except ValueError:
            continue

    timestamp_match = re.search(r"\b(1\d{9,12})\b", text)
    if timestamp_match:
        try:
            raw_ts = int(timestamp_match.group(1))
            if raw_ts > 10_000_000_000:
                raw_ts = raw_ts / 1000
            return datetime.fromtimestamp(raw_ts)
        except Exception:
            return None

    return None


def profile_content_visible(page, timeout_ms: int = 500) -> bool:
    return wait_for_selector(page, PAGE_READY_SELECTOR, timeout_ms)


def page_shell_visible(page, timeout_ms: int = 500) -> bool:
    return wait_for_selector(page, PAGE_SHELL_READY_SELECTOR, timeout_ms)


def visible_post_anchor_count(page) -> int:
    try:
        return int(page.locator(POST_LINK_SELECTOR).count())
    except Exception:
        return 0


def has_authenticated_session(context) -> bool:
    try:
        cookies = context.cookies()
    except Exception:
        return False
    return any(cookie.get("name") == "c_user" for cookie in cookies)


def url_indicates_checkpoint_or_verification(url: str) -> bool:
    normalized = (url or "").lower()
    return any(token in normalized for token in VERIFICATION_URL_TOKENS)


def current_url_matches_target(current_url: str, target_url: str) -> bool:
    if not current_url or not target_url:
        return False
    try:
        current = urlparse(current_url)
        target = urlparse(target_url)
    except Exception:
        return False
    current_host = (current.hostname or "").lower()
    target_host = (target.hostname or "").lower()
    if current_host != target_host:
        return False
    current_path = (current.path or "/").rstrip("/") or "/"
    target_path = (target.path or "/").rstrip("/") or "/"
    return current_path == target_path


def detect_checkpoint_or_verification(page) -> tuple[bool, str]:
    current_url = ""
    try:
        current_url = page.url or ""
    except Exception:
        current_url = ""

    if url_indicates_checkpoint_or_verification(current_url):
        return True, "Facebook checkpoint or verification page detected."

    try:
        body_text = page.locator("body").inner_text(timeout=600).lower()
    except Exception:
        body_text = ""

    challenge_phrases = [
        "confirm your identity",
        "check your notifications on another device",
        "enter the code we sent",
        "approve from another device",
        "security check",
        "verification required",
        "suspicious login attempt",
        "suspicious login",
        "checkpoint",
        "verify your account",
        "review recent login",
        "recaptcha",
        "captcha",
        "google's recaptcha enterprise",
        "protect the security of your account",
        "authentication required",
    ]
    if any(phrase in body_text for phrase in challenge_phrases):
        return True, "Facebook requires verification before the page can be used."

    return False, ""


def detect_login_gate(page) -> tuple[bool, str]:
    if wait_for_selector(page, LOGIN_FORM_SELECTOR, 350):
        return True, "Facebook login form detected."

    blocked, blocked_reason = detect_checkpoint_or_verification(page)
    if blocked:
        return True, blocked_reason

    current_url = ""
    try:
        current_url = page.url or ""
    except Exception:
        current_url = ""

    if any(token in current_url for token in ["/login", "/checkpoint", "/recover"]):
        return True, "Facebook login or checkpoint page detected."

    try:
        body_text = page.locator("body").inner_text(timeout=600).lower()
    except Exception:
        body_text = ""

    blocking_phrases = [
        "log in to continue",
        "you must log in",
        "see more posts from",
        "continue on facebook",
        "create new account",
        "continue reading this story by logging in",
        "see more from",
        "log in to see more",
    ]
    if any(phrase in body_text for phrase in blocking_phrases):
        return True, "Facebook login wall detected."

    login_cta_visible = False
    try:
        login_cta_visible = page.locator(
            "a[href*='/login'], a[href*='login.php'], button[name='login'], form[action*='login'] button"
        ).count() > 0
    except Exception:
        login_cta_visible = False

    if login_cta_visible and ("log in" in body_text or "sign up" in body_text):
        return True, "Facebook requires login before deeper content can be collected."

    if ("log in" in body_text and "sign up" in body_text) and visible_post_anchor_count(page) < 5:
        return True, "Facebook public gate detected."

    return False, ""


def facebook_strong_ready_signal(page, target_url: str = "") -> tuple[bool, str]:
    checkpoint_required, checkpoint_reason = detect_checkpoint_or_verification(page)
    if checkpoint_required:
        return False, checkpoint_reason

    login_required, login_reason = detect_login_gate(page)
    if login_required:
        return False, login_reason

    signals = [
        (POST_LINK_SELECTOR, "post links visible"),
        (FEED_READY_SELECTOR, "feed/posts container visible"),
        ("[aria-label='Account'], div[aria-label='Account']", "account menu visible"),
        ("a[aria-label='Message'], div[aria-label='Message'], span:has-text('Message')", "message button visible"),
        (PAGE_SHELL_READY_SELECTOR, "page shell/profile content visible"),
        ("div[role='navigation']", "navigation bar visible"),
        ("div[role='main']", "main content visible"),
        ("svg[aria-label='Your profile'], [aria-label='Your profile']", "user avatar/nav visible"),
    ]

    for selector, reason in signals:
        if wait_for_selector(page, selector, 700):
            try:
                state = get_scroll_state(page)
                if state.get("visiblePosts", 0) > 0:
                    return True, reason
            except Exception:
                return True, reason

    try:
        current_url = page.url or ""
    except Exception:
        current_url = ""

    if target_url and current_url_matches_target(current_url, target_url):
        if not wait_for_selector(page, LOGIN_FORM_SELECTOR, 300):
            return True, "target URL loaded and login form is gone"

    return False, "Facebook page is not ready yet."


def page_ready_for_collection(page, target_url: str = "") -> bool:
    ready, _ = facebook_strong_ready_signal(page, target_url)
    return ready


def validate_session(page, context=None, target_url: str = "") -> dict[str, Any]:
    checkpoint_required, checkpoint_reason = detect_checkpoint_or_verification(page)
    if checkpoint_required:
        return {
            "state": "verification_required",
            "reason": checkpoint_reason,
            "url": getattr(page, "url", target_url) or target_url,
            "cookiesPresent": bool(context and has_authenticated_session(context)),
        }

    login_required, login_reason = detect_login_gate(page)
    if login_required:
        return {
            "state": "login_required",
            "reason": login_reason,
            "url": getattr(page, "url", target_url) or target_url,
            "cookiesPresent": bool(context and has_authenticated_session(context)),
        }

    ready, ready_reason = facebook_strong_ready_signal(page, target_url)
    if ready:
        return {
            "state": "ready",
            "reason": f"Facebook readiness check passed: {ready_reason}.",
            "url": getattr(page, "url", target_url) or target_url,
            "cookiesPresent": bool(context and has_authenticated_session(context)),
        }

    return {
        "state": "unknown",
        "reason": "Facebook session could not be confirmed yet.",
        "url": getattr(page, "url", target_url) or target_url,
        "cookiesPresent": bool(context and has_authenticated_session(context)),
    }


def manual_login_url(target_url: str) -> str:
    return f"https://www.facebook.com/login.php?next={quote(target_url, safe='')}"


def click_button_if_visible(page, pattern: str, timeout_ms: int = 1500) -> bool:
    try:
        button = page.get_by_role("button", name=re.compile(pattern, re.IGNORECASE)).first
        if button.count() > 0:
            button.click(timeout=timeout_ms)
            return True
    except Exception:
        pass
    return False


def save_storage_state(context, log_hook: Optional[LogHook] = None) -> None:
    path = get_storage_state_path()
    if path is None:
        return
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        try:
            context.storage_state(path=str(path), indexed_db=True)
        except TypeError:
            context.storage_state(path=str(path))
        emit_log(log_hook, "INFO", "Storage state saved", str(path))
    except Exception as exc:
        emit_log(log_hook, "WARN", "Session save skipped", type(exc).__name__)


def auto_login_if_needed(page, context, target_url: str, log_hook: Optional[LogHook] = None) -> bool:
    if not has_login_credentials():
        return False
    if not wait_for_selector(page, LOGIN_FORM_SELECTOR, LOGIN_FORM_TIMEOUT):
        return False

    emit_log(log_hook, "INFO", "Facebook login", "Login form detected. Attempting env-based sign-in.")
    try:
        page.locator("input[name='email']").first.fill(FACEBOOK_USERNAME)
        page.locator("input[name='pass']").first.fill(FACEBOOK_PASSWORD)
        if page.locator("button[name='login']").count() > 0:
            page.locator("button[name='login']").first.click(timeout=3000)
        else:
            page.locator("button[type='submit']").first.click(timeout=3000)
    except Exception as exc:
        emit_log(log_hook, "WARN", "Facebook login failed", type(exc).__name__)
        return False

    try:
        page.wait_for_load_state("domcontentloaded", timeout=LOGIN_POST_SUBMIT_TIMEOUT)
    except Exception:
        pass

    click_button_if_visible(page, r"not now|skip|maybe later")
    page.goto(target_url, wait_until="domcontentloaded", timeout=POST_GOTO_TIMEOUT)
    apply_local_page_preferences(page)

    if page_ready_for_collection(page) and has_authenticated_session(context):
        save_storage_state(context, log_hook)
        emit_log(log_hook, "SUCCESS", "Facebook login", "Session is ready and the target page is visible.")
        return True

    emit_log(log_hook, "WARN", "Facebook login incomplete", "Target page did not become ready after sign-in.")
    return False


def route_nonessential_resources(route) -> None:
    if route.request.resource_type in BLOCKED_RESOURCE_TYPES:
        route.abort()
        return
    route.continue_()


def install_playwright_browsers() -> None:
    command = [sys.executable, "-m", "playwright", "install", "chromium", "chromium-headless-shell"]
    subprocess.run(command, check=True, timeout=240)


def launch_browser(playwright):
    global ACTIVE_BROWSER_ENGINE
    headless = False if uses_local_browser_window() else PLAYWRIGHT_HEADLESS
    launch_options = {
        "headless": headless,
        "args": [
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-dev-shm-usage",
            "--disable-notifications",
            "--deny-permission-prompts",
            "--disable-features=NotificationTriggers,PermissionsPromptService,PushMessaging",
            "--disable-blink-features=AutomationControlled",
            "--start-maximized",
            "--no-first-run",
            "--no-default-browser-check",
        ],
    }
    if uses_local_browser_window():
        user_data_dir = get_user_data_dir()
        if user_data_dir is None:
            raise RuntimeError("Facebook local browser mode requires a user data directory.")
        user_data_dir.mkdir(parents=True, exist_ok=True)
        persistent_options = {
            "headless": False,
            "viewport": {"width": 1920, "height": 1080},
            "locale": "en-US",
            "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "args": launch_options["args"],
        }
        if PLAYWRIGHT_BROWSER_CHANNEL:
            persistent_options["channel"] = PLAYWRIGHT_BROWSER_CHANNEL
        try:
            context = playwright.chromium.launch_persistent_context(
                user_data_dir=str(user_data_dir),
                **persistent_options,
            )
            apply_context_preferences(context)
            ACTIVE_BROWSER_ENGINE = PLAYWRIGHT_BROWSER_CHANNEL or "chromium"
        except PlaywrightError as exc:
            message = str(exc)
            missing_browser = "Executable doesn't exist" in message or "Please run the following command" in message
            if PLAYWRIGHT_BROWSER_CHANNEL:
                fallback_options = dict(persistent_options)
                fallback_options.pop("channel", None)
                try:
                    context = playwright.chromium.launch_persistent_context(
                        user_data_dir=str(user_data_dir),
                        **fallback_options,
                    )
                    apply_context_preferences(context)
                    ACTIVE_BROWSER_ENGINE = "chromium"
                    return None, context
                except PlaywrightError:
                    pass
            if not (PLAYWRIGHT_AUTO_INSTALL and missing_browser):
                raise
            install_playwright_browsers()
            fallback_options = dict(persistent_options)
            fallback_options.pop("channel", None)
            context = playwright.chromium.launch_persistent_context(
                user_data_dir=str(user_data_dir),
                **fallback_options,
            )
            apply_context_preferences(context)
            ACTIVE_BROWSER_ENGINE = "chromium"
        return None, context

    try:
        browser = playwright.chromium.launch(**launch_options)
        ACTIVE_BROWSER_ENGINE = "chromium"
    except PlaywrightError as exc:
        message = str(exc)
        missing_browser = "Executable doesn't exist" in message or "Please run the following command" in message
        if not (PLAYWRIGHT_AUTO_INSTALL and missing_browser):
            raise
        install_playwright_browsers()
        browser = playwright.chromium.launch(**launch_options)
        ACTIVE_BROWSER_ENGINE = "chromium"

    context, _ = load_or_create_context(browser)
    return browser, context


def apply_local_page_preferences(page) -> None:
    if not uses_local_browser_window():
        return
    try:
        # Reset zoom to 100%
        page.keyboard.press("Control+0")
        page.wait_for_timeout(100)
    except Exception:
        try:
            page.evaluate("document.body.style.zoom = '100%';")
        except Exception:
            pass


def normalize_facebook_page_viewport(page, log_hook: Optional[LogHook] = None) -> None:
    """Normalize viewport, zoom, and scroll position for stable Facebook content positioning."""
    try:
        # Reset zoom to 100%
        page.keyboard.press("Control+0")
        page.wait_for_timeout(150)
    except Exception:
        try:
            page.evaluate(
                """() => {
                    for (const elem of [document.documentElement, document.body]) {
                        if (elem) {
                            elem.style.zoom = '100%';
                        }
                    }
                }"""
            )
        except Exception:
            pass

    # Scroll to top before starting collection
    try:
        page.evaluate("window.scrollTo(0, 0);")
        page.wait_for_timeout(300)
    except Exception:
        pass

    # Log viewport normalization
    if log_hook:
        try:
            viewport = page.evaluate("() => ({ width: window.innerWidth, height: window.innerHeight, scrollY: window.scrollY })")
            emit_log(log_hook, "INFO", "Viewport normalized", f"Position: {viewport}")
        except Exception:
            emit_log(log_hook, "INFO", "Viewport normalized", "Page positioning reset to top with 100% zoom")


def focus_posts_section(page, log_hook: Optional[LogHook] = None) -> None:
    script = """
        () => {
            const isVisible = (node) => {
                if (!node) return false;
                const style = window.getComputedStyle(node);
                if (!style || style.visibility === 'hidden' || style.display === 'none') return false;
                const rect = node.getBoundingClientRect();
                return rect.width > 8 && rect.height > 8;
            };

            const candidates = Array.from(document.querySelectorAll("h1, h2, h3, span, div"));
            for (const node of candidates) {
                const text = (node.textContent || '').trim().toLowerCase();
                if (text !== 'posts') continue;
                if (!isVisible(node)) continue;
                node.scrollIntoView({ block: 'start', inline: 'nearest' });
                return true;
            }
            return false;
        }
    """
    try:
        moved = bool(page.evaluate(script))
        if moved:
            page.wait_for_timeout(250)
            emit_log(log_hook, "INFO", "Posts section", "Scrolled the Facebook page to the posts section before collecting links.")
        try:
            page.wait_for_load_state("networkidle", timeout=8000)
        except Exception:
            pass
        page.wait_for_timeout(1200)
    except Exception:
        return


def collect_visible_post_links(page, target_url: str = "") -> list[str]:
    return page.evaluate(
        """(targetUrl) => {
            const normalizeHref = (value) => {
                if (!value) return '';
                try {
                    const url = new URL(value, window.location.origin);
                    for (const key of ['comment_id', 'reply_comment_id', '__tn__', '__cft__', 'ref', 'refsrc', 'notif_t', 'comment_tracking', 'acontext']) {
                        url.searchParams.delete(key);
                    }
                    return `${url.origin}${url.pathname}${url.search}`;
                } catch (error) {
                    return String(value || '').split('&__')[0];
                }
            };
            const isVisible = (node) => {
                if (!node) return false;
                const style = window.getComputedStyle(node);
                if (!style || style.visibility === 'hidden' || style.display === 'none') return false;
                const rect = node.getBoundingClientRect();
                return rect.width > 8 && rect.height > 8;
            };
            const target = normalizeHref(targetUrl);
            let targetSlug = '';
            try {
                if (target) {
                    const parsed = new URL(target);
                    targetSlug = (parsed.pathname || '').split('/').filter(Boolean)[0] || '';
                }
            } catch (error) {}
            const selectors = [
                "a[href*='/posts/']",
                "a[href*='/videos/']",
                "a[href*='/permalink/']",
                "a[href*='story_fbid=']",
                "a[href*='/photo.php']",
                "a[href*='/watch/?v=']",
            ];
            const links = [];
            const articleSelectors = [
                "div[role='main'] div[role='article']",
                "div[role='feed'] div[role='article']",
                "div[data-pagelet*='FeedUnit']",
                "div[aria-posinset]",
            ];
            const articleNodes = new Set();
            for (const selector of articleSelectors) {
                for (const node of document.querySelectorAll(selector)) {
                    if (!isVisible(node)) continue;
                    const text = (node.innerText || '').trim().toLowerCase();
                    if (!text) continue;
                    if (
                        text.includes('people you may know') ||
                        text.includes('suggested for you') ||
                        text.includes('sponsored')
                    ) continue;
                    const inFeatured = Boolean(
                        node.closest("[aria-label='Featured'], [data-pagelet*='Featured'], section, div")
                        && /featured/i.test(node.closest("section, div")?.innerText || '')
                    );
                    if (inFeatured) continue;
                    articleNodes.add(node);
                }
            }
            const roots = articleNodes.size ? Array.from(articleNodes) : [document.querySelector("div[role='main']") || document];
            for (const root of roots) {
                for (const selector of selectors) {
                    for (const anchor of root.querySelectorAll(selector)) {
                        const href = normalizeHref(anchor.href || "");
                        if (!href) continue;
                        if (targetSlug && !href.includes(`/${targetSlug}/`) && !href.includes('story_fbid=')) {
                            // Allow broader matches; dedupe + normalization will filter out unrelated items.
                        }
                        links.push(href);
                    }
                }
            }
            return links;
        }""",
        arg=target_url,
    )


def get_scroll_state(page) -> dict[str, Any]:
    return page.evaluate(
        """() => {
            const doc = document.scrollingElement || document.documentElement;
            const feed = document.querySelector("div[role='feed']");
            const visiblePosts = document.querySelectorAll(
                "div[role='main'] div[role='article'], div[role='feed'] div[role='article'], div[data-pagelet*='FeedUnit'], div[aria-posinset]"
            ).length;
            const anchors = new Set(
                Array.from(document.querySelectorAll("a[href*='/posts/'], a[href*='/videos/'], a[href*='/permalink/'], a[href*='story_fbid='], a[href*='/photo.php'], a[href*='/watch/?v=']")).map(a => (a.href || '').split('&__')[0]).filter(Boolean)
            );
            const articleCount = document.querySelectorAll("div[role='main'] div[role='article'], div[role='feed'] div[role='article'], div[data-pagelet*='FeedUnit'], div[aria-posinset]").length;
            const bodyHeight = Math.max(
                Number(doc.scrollHeight || 0),
                Number(document.body ? document.body.scrollHeight : 0),
                Number(document.documentElement ? document.documentElement.scrollHeight : 0)
            );
            const top = Number(doc.scrollTop || window.pageYOffset || 0);
            const viewport = Number(window.innerHeight || doc.clientHeight || 0);
            return {
                feedDetected: Boolean(feed),
                visiblePosts,
                linkCount: anchors.size,
                articleCount,
                scrollTop: top,
                scrollHeight: bodyHeight,
                bodyHeight,
                atBottom: top + viewport >= bodyHeight - 32,
            };
        }"""
    )


def apply_scroll_strategy(page, strategy: str) -> None:
    """Apply scroll strategy with deterministic scroll distances.
    
    Args:
        strategy: Scroll method to use (window-scroll, mouse-wheel, page-down, bottom-jump)
    """
    if strategy == "window-scroll":
        # Use 75% of viewport height for consistent scrolling (typically ~675px at 900px height)
        page.evaluate("window.scrollBy(0, Math.max(window.innerHeight * 0.75, 550));")
    elif strategy == "mouse-wheel":
        # Mouse wheel: 2200px per wheel units
        page.mouse.wheel(0, 2200)
    elif strategy == "page-down":
        page.keyboard.press("PageDown")
    else:
        # bottom-jump: scroll to absolute bottom
        page.evaluate("window.scrollTo(0, document.body.scrollHeight);")


def wait_for_scroll_stabilization(page, timeout_ms: int = SCROLL_WAIT_TIMEOUT) -> dict[str, Any]:
    """
    REAL scroll stabilization: Wait until scrolling stops and DOM settles.
    
    Waits for:
    - scrollY to stop changing (2 consecutive checks with same value)
    - document.body.scrollHeight to stabilize
    - post elements count to stabilize for 2 cycles
    """
    try:
        page.wait_for_function(
            """(timeout_ms) => {
                const start = Date.now();
                let lastScrollY = window.scrollY;
                let lastScrollHeight = document.body.scrollHeight;
                let lastPostCount = document.querySelectorAll("a[href*='/posts/'], a[href*='/videos/'], a[href*='/permalink/'], a[href*='story_fbid='], a[href*='/photo.php']").length;
                let scrollStableCount = 0;
                let heightStableCount = 0;
                let postCountStableCount = 0;
                
                return new Promise(resolve => {
                    const checkStability = () => {
                        const now = Date.now();
                        if (now - start > timeout_ms) {
                            resolve(true);
                            return;
                        }
                        
                        const currentScrollY = window.scrollY;
                        const currentScrollHeight = document.body.scrollHeight;
                        const currentPostCount = document.querySelectorAll("a[href*='/posts/'], a[href*='/videos/'], a[href*='/permalink/'], a[href*='story_fbid='], a[href*='/photo.php']").length;
                        
                        // Check stability
                        if (Math.abs(currentScrollY - lastScrollY) < 5) {
                            scrollStableCount++;
                        } else {
                            scrollStableCount = 0;
                        }
                        
                        if (Math.abs(currentScrollHeight - lastScrollHeight) < 5) {
                            heightStableCount++;
                        } else {
                            heightStableCount = 0;
                        }
                        
                        if (Math.abs(currentPostCount - lastPostCount) === 0) {
                            postCountStableCount++;
                        } else {
                            postCountStableCount = 0;
                        }
                        
                        lastScrollY = currentScrollY;
                        lastScrollHeight = currentScrollHeight;
                        lastPostCount = currentPostCount;
                        
                        // All stable for 2 cycles = DOM is ready
                        if (scrollStableCount >= 2 && heightStableCount >= 2 && postCountStableCount >= 2) {
                            resolve(true);
                        } else {
                            setTimeout(checkStability, 150);
                        }
                    };
                    checkStability();
                });
            }""",
            arg=timeout_ms,
            timeout=timeout_ms + 1000,
        )
    except Exception:
        pass
    return get_scroll_state(page)


def wait_for_scroll_growth(page, previous_state: dict[str, Any], timeout_ms: int = SCROLL_WAIT_TIMEOUT) -> dict[str, Any]:
    """Original scroll growth check, now called after stabilization."""
    try:
        page.wait_for_function(
            """(prev) => {
                const doc = document.scrollingElement || document.documentElement;
                const anchors = new Set(
                    Array.from(document.querySelectorAll("a[href*='/posts/'], a[href*='/videos/'], a[href*='/permalink/'], a[href*='story_fbid='], a[href*='/photo.php'], a[href*='/watch/?v=']")).map(a => (a.href || '').split('&__')[0]).filter(Boolean)
                );
                const articleCount = document.querySelectorAll("div[role='main'] div[role='article'], div[role='feed'] div[role='article'], div[data-pagelet*='FeedUnit'], div[aria-posinset]").length;
                const bodyHeight = Math.max(
                    Number(doc.scrollHeight || 0),
                    Number(document.body ? document.body.scrollHeight : 0),
                    Number(document.documentElement ? document.documentElement.scrollHeight : 0)
                );
                const top = Number(doc.scrollTop || window.pageYOffset || 0);
                return anchors.size > prev.linkCount || articleCount > prev.articleCount || bodyHeight > prev.bodyHeight + 24 || top > prev.scrollTop + 24;
            }""",
            arg=previous_state,
            timeout=timeout_ms,
        )
    except Exception:
        pass
    return get_scroll_state(page)


def validate_facebook_feed_ready(page, target_url: str = "", log_hook: Optional[LogHook] = None) -> tuple[bool, str]:
    """
    STRICT feed validation - DO NOT start collection unless feed is confirmed ready.
    
    Checks:
    - Feed container is visible
    - No loading skeletons
    - No spinners
    - At least 1 post card present
    - DOM is stable
    """
    try:
        # Check feed container exists and is visible
        result = page.evaluate("""() => {
            const feedContainers = [
                document.querySelector("div[role='feed']"),
                document.querySelector("div[role='main']"),
                document.querySelector("div[data-pagelet*='Feed']"),
            ].filter(Boolean);
            
            if (!feedContainers.length) return { found: false, reason: "Feed container not found" };
            
            const feed = feedContainers[0];
            const style = window.getComputedStyle(feed);
            if (style.display === 'none' || style.visibility === 'hidden') {
                return { found: false, reason: "Feed container is hidden" };
            }
            
            // Check for loading indicators
            const loadingSelectors = [
                'div[role="status"]',  // WAI status/loading
                'div[aria-label*="load"]',
                '[data-testid*="loading"]',
                'div:has(> svg):has(> circle)', // Spinner
            ];
            
            for (const selector of loadingSelectors) {
                const loader = document.querySelector(selector);
                if (loader) {
                    const style = window.getComputedStyle(loader);
                    if (style.display !== 'none' && style.visibility !== 'hidden') {
                        return { found: false, reason: `Loading indicator detected: ${selector}` };
                    }
                }
            }
            
            // Check for skeleton loaders
            const skeletons = document.querySelectorAll('[aria-label*="skeleton"], [data-testid*="skeleton"]');
            if (skeletons.length > 0) {
                const visibleSkeletons = Array.from(skeletons).filter(s => {
                    const style = window.getComputedStyle(s);
                    return style.display !== 'none' && style.visibility !== 'hidden';
                });
                if (visibleSkeletons.length > 0) {
                    return { found: false, reason: `${visibleSkeletons.length} skeleton loaders still visible` };
                }
            }
            
            // Check for actual post cards
            const postSelectors = [
                'a[href*="/posts/"]',
                'a[href*="/videos/"]',
                'a[href*="/permalink/"]',
                'a[href*="story_fbid="]',
                'a[href*="/photo.php"]',
            ];
            
            let postCount = 0;
            for (const selector of postSelectors) {
                postCount += document.querySelectorAll(selector).length;
            }
            
            if (postCount === 0) {
                return { found: false, reason: "No post cards found in feed" };
            }
            
            return { found: true, reason: `Feed ready with ${postCount} post cards`, postCount };
        }""")
        
        if result.get("found"):
            emit_log(log_hook, "SUCCESS", "Feed validation", result.get("reason"))
            return True, result.get("reason", "Feed ready")
        else:
            emit_log(log_hook, "WARN", "Feed not ready", result.get("reason", "Unknown reason"))
            return False, result.get("reason", "Feed not ready")
            
    except Exception as e:
        emit_log(log_hook, "ERROR", "Feed validation failed", str(e))
        return False, f"Feed validation error: {str(e)}"


def collect_post_links(
    page,
    scroll_rounds: int,
    target_url: str = "",
    log_hook: Optional[LogHook] = None,
    progress_hook: Optional[ProgressHook] = None,
    cancel_check: Optional[Callable[[], bool]] = None,
    diagnostics: Optional[dict[str, Any]] = None,
) -> list[str]:
    links: dict[str, bool] = {}
    stagnant_rounds = 0
    stagnant_limit = max(8, min(18, max(8, scroll_rounds))) if scroll_rounds > 0 else 8
    min_rounds_before_stop = max(6, min(scroll_rounds, max(6, scroll_rounds // 2))) if scroll_rounds > 0 else 6
    strategies = ("window-scroll", "mouse-wheel", "page-down", "bottom-jump")
    stop_reason = f"Reached max scroll rounds ({scroll_rounds})."

    if not target_url:
        try:
            target_url = page.url or ""
        except Exception:
            target_url = ""

    focus_posts_section(page, log_hook=log_hook)

    # Normalize page viewport and zoom for stable scrolling
    normalize_facebook_page_viewport(page, log_hook=log_hook)

    emit_log(log_hook, "INFO", "Checking login state", "Verifying Facebook access before scrolling.")
    
    # STRICT FEED VALIDATION - DO NOT PROCEED WITHOUT CONFIRMED FEED
    feed_ready, feed_reason = validate_facebook_feed_ready(page, target_url=target_url, log_hook=log_hook)
    if not feed_ready:
        emit_log(log_hook, "WARN", "Feed not ready", f"Retrying feed validation...")
        page.wait_for_timeout(500)
        feed_ready, feed_reason = validate_facebook_feed_ready(page, target_url=target_url, log_hook=log_hook)
    
    if not feed_ready:
        emit_log(log_hook, "ERROR", "Feed validation failed", f"Cannot start collection: {feed_reason}")
        return []
    
    initial_state = get_scroll_state(page)
    emit_log(
        log_hook,
        "INFO",
        "Visibility debug",
        (
            f"Feed detected: {'YES' if initial_state.get('feedDetected') else 'NO'}, "
            f"visible posts: {initial_state.get('visiblePosts', 0)}, anchors: {initial_state.get('linkCount', 0)}."
        ),
    )

    initial_links = dedupe_post_links(collect_visible_post_links(page, target_url=target_url), target_url=target_url)
    for href in initial_links:
        if href and href not in links:
            links[href] = True

    emit_log(log_hook, "INFO", "Initial content", f"+{len(initial_links)} links visible (total={len(links)}).")
    emit_progress(progress_hook, 0, scroll_rounds, len(links))

    for round_index in range(1, scroll_rounds + 1):
        if cancel_check and cancel_check():
            raise RuntimeError("Cancelled during Facebook scrolling.")

        checkpoint_required, checkpoint_reason = detect_checkpoint_or_verification(page)
        if checkpoint_required:
            emit_log(log_hook, "WARN", "Facebook checkpoint", checkpoint_reason)
            raise AuthRequiredError("waiting_verification", checkpoint_reason)

        login_required, login_reason = detect_login_gate(page)
        if login_required:
            emit_log(log_hook, "WARN", "Facebook login required", login_reason)
            raise AuthRequiredError("waiting_login", login_reason)

        round_start = time.perf_counter()
        before_state = get_scroll_state(page)
        emit_log(
            log_hook,
            "INFO",
            "Visibility debug",
            (
                f"Round {round_index}: feed={'YES' if before_state.get('feedDetected') else 'NO'}, "
                f"visible posts={before_state.get('visiblePosts', 0)}, anchors={before_state.get('linkCount', 0)}."
            ),
        )
        before_count = len(links)
        strategy_used = "none"
        height_before = before_state["bodyHeight"]
        height_after = height_before
        anchors_after = before_state["linkCount"]
        article_count_after = before_state.get("articleCount", 0)

        for attempt_index, strategy in enumerate(strategies, start=1):
            strategy_used = strategy
            emit_log(log_hook, "INFO", f"Scroll {round_index}", f"Attempt {attempt_index}: {strategy}.")
            apply_scroll_strategy(page, strategy)
            try:
                page.mouse.move(120 + attempt_index * 10, 360 + attempt_index * 6, steps=5)
            except Exception:
                pass
            page.wait_for_timeout(200)
            
            # REAL STABILIZATION: Wait for DOM to truly settle before checking growth
            after_state = wait_for_scroll_stabilization(page, timeout_ms=SCROLL_WAIT_TIMEOUT)
            height_after = after_state["bodyHeight"]
            anchors_after = after_state["linkCount"]
            article_count_after = after_state.get("articleCount", before_state.get("articleCount", 0))

            fresh_links = dedupe_post_links(collect_visible_post_links(page, target_url=target_url), target_url=target_url)
            for href in fresh_links:
                if href and href not in links:
                    links[href] = True

            if (
                len(links) > before_count
                or after_state.get("articleCount", 0) > before_state.get("articleCount", 0)
                or after_state["bodyHeight"] > before_state["bodyHeight"] + 24
                or after_state["scrollTop"] > before_state["scrollTop"] + 24
            ):
                break

            page.wait_for_timeout(PROFILE_RETRY_MS + 400)

        new_links = len(links) - before_count
        article_growth = max(0, article_count_after - before_state.get("articleCount", 0))
        elapsed = time.perf_counter() - round_start
        if new_links == 0 and article_growth == 0:
            stagnant_rounds += 1
            detail = (
                f"+0 new links (total={len(links)}, strategy={strategy_used}, "
                f"height={height_before}->{height_after}, anchors={before_state['linkCount']}->{anchors_after}, "
                f"articles={before_state.get('articleCount', 0)}->{article_count_after}, "
                f"stagnant={stagnant_rounds}/{stagnant_limit})"
            )
            if before_state["atBottom"]:
                detail += ", bottom reached"
            emit_log(log_hook, "INFO", f"Scroll {round_index}", detail)
        else:
            stagnant_rounds = 0
            detail_suffix = f"+{new_links} new links" if new_links > 0 else "+0 new links"
            if article_growth > 0:
                detail_suffix += f", +{article_growth} new post containers"
            emit_log(
                log_hook,
                "INFO",
                f"Scroll {round_index}",
                (
                    f"{detail_suffix} (total={len(links)}, strategy={strategy_used}, "
                    f"height={height_before}->{height_after}, anchors={before_state['linkCount']}->{anchors_after}, "
                    f"articles={before_state.get('articleCount', 0)}->{article_count_after})"
                ),
            )

        if elapsed >= SLOW_SCROLL_SECONDS:
            emit_log(log_hook, "WARN", "Slow scroll", f"Round {round_index} took {elapsed:.2f}s.")

        emit_progress(progress_hook, round_index, scroll_rounds, len(links))

        if stagnant_rounds >= stagnant_limit and round_index >= min_rounds_before_stop:
            stop_reason = f"Confirmed stagnation after {stagnant_rounds} rounds with no new Facebook links."
            break

    if diagnostics is not None:
        diagnostics["stopReason"] = stop_reason
        diagnostics["totalLinks"] = len(links)

    if not links:
        emit_log(log_hook, "WARN", "No links collected", "No Facebook post links were found after scrolling. Stopping gracefully.")

    emit_log(log_hook, "INFO", "Link collection stop", stop_reason)
    return list(links.keys())


def infer_post_type(url: str) -> str:
    lower_url = url.lower()
    if "/videos/" in lower_url or "/watch/" in lower_url:
        return "Video"
    if "/photo.php" in lower_url:
        return "Photo"
    if "story_fbid=" in lower_url:
        return "Story / Post"
    return "Post"


def summarize_log_text(text: str, limit: int = 120) -> str:
    compact = re.sub(r"\s+", " ", text or "").strip()
    if len(compact) <= limit:
        return compact
    return f"{compact[:limit - 3]}..."


def inspect_active_post_scope(page, target_url: str = "") -> dict[str, Any]:
    script = """
        (targetUrl) => {
            const normalizeHref = (value) => {
                if (!value) return '';
                try {
                    const url = new URL(value, window.location.origin);
                    for (const key of ['comment_id', 'reply_comment_id', '__tn__', '__cft__', 'ref', 'refsrc', 'notif_t', 'comment_tracking', 'acontext']) {
                        url.searchParams.delete(key);
                    }
                    return `${url.origin}${url.pathname}${url.search}`;
                } catch (error) {
                    return String(value || '').split('&__')[0];
                }
            };

            const isVisible = (node) => {
                if (!node) return false;
                const style = window.getComputedStyle(node);
                if (!style || style.visibility === 'hidden' || style.display === 'none') return false;
                const rect = node.getBoundingClientRect();
                return rect.width > 8 && rect.height > 8;
            };

            const target = normalizeHref(targetUrl);
            let targetPath = '';
            let targetSlug = '';
            try {
                if (target) {
                    const parsed = new URL(target);
                    targetPath = `${parsed.pathname}${parsed.search}`;
                    targetSlug = (parsed.pathname || '').split('/').filter(Boolean)[0] || '';
                }
            } catch (error) {}

            const candidates = [];
            const addCandidate = (node, type, baseScore) => {
                if (!node || !isVisible(node)) return;
                const text = (node.innerText || '').trim();
                if (!text) return;
                candidates.push({ node, type, baseScore, text });
            };

            for (const dialog of document.querySelectorAll("div[role='dialog']")) {
                addCandidate(dialog, 'dialog', 120);
            }

            const articleSelectors = [
                "div[role='main'] div[role='article']",
                "div[role='feed'] div[role='article']",
                "div[data-pagelet*='FeedUnit']",
                "div[aria-posinset]",
            ];
            for (const selector of articleSelectors) {
                for (const article of document.querySelectorAll(selector)) {
                    addCandidate(article, 'article', 30);
                }
            }

            if (!candidates.length) {
                const main = document.querySelector("div[role='main']");
                addCandidate(main, 'main', 10);
            }

            let best = null;
            for (const candidate of candidates) {
                const { node, type, baseScore, text } = candidate;
                const links = Array.from(node.querySelectorAll("a[href]"))
                    .map((anchor) => normalizeHref(anchor.href))
                    .filter(Boolean);
                const matchedTarget = Boolean(
                    target &&
                    links.some((href) => href === target || (targetPath && href.includes(targetPath)))
                );
                const matchedSlug = Boolean(
                    targetSlug &&
                    links.some((href) => href.includes(`/${targetSlug}/`) || href.includes(`/${targetSlug}?`))
                );
                let score = baseScore;
                if (matchedTarget) score += 160;
                if (matchedSlug) score += 45;
                if (node.querySelector("abbr, time")) score += 35;
                if (links.length) score += 10;
                if (/(reactions?|comments?|shares?|likes?)/i.test(text)) score += 18;
                if (/featured/i.test(text) && !matchedTarget) score -= 120;
                if (/people you may know|suggested for you|sponsored/i.test(text)) score -= 180;
                score += Math.min(30, Math.floor(text.length / 160));
                if (!best || score > best.score) {
                    best = { node, type, text, matchedTarget, matchedSlug, score };
                }
            }

            if (!best) {
                return {
                    found: false,
                    scopeType: '',
                    matchedTarget: false,
                    matchedSlug: false,
                    scopeText: '',
                    metricTexts: [],
                    actionTexts: [],
                    dateCandidates: [],
                    preview: '',
                };
            }

            const root = best.node;
            const rootText = (root.innerText || '').replace(/\\s+/g, ' ').trim();
            const metricTexts = [];
            const seenMetricTexts = new Set();
            const actionTexts = [];
            const seenActionTexts = new Set();
            const metricSelectors = [
                "[aria-label]",
                "[title]",
                "div[role='button']",
                "button",
                "a[href]",
                "a[role='link']",
                "span",
                "div",
            ];
            for (const selector of metricSelectors) {
                for (const node of root.querySelectorAll(selector)) {
                    if (!isVisible(node)) continue;
                    const rawParts = [
                        node.getAttribute('aria-label') || '',
                        node.getAttribute('title') || '',
                        (node.textContent || '').trim(),
                    ].filter(Boolean);
                    for (const rawPart of rawParts) {
                        const cleaned = rawPart.replace(/\\s+/g, ' ').trim();
                        if (!cleaned || cleaned.length > 160 || seenMetricTexts.has(cleaned)) continue;
                        if (/[0-9]/.test(cleaned) || /(reaction|like|comment|share)/i.test(cleaned)) {
                            seenMetricTexts.add(cleaned);
                            metricTexts.push(cleaned);
                        }
                    }
                }
            }

            const actionSelectors = [
                "a[href*='comment_id=']",
                "a[href*='/ufi/reaction/profile/browser/']",
                "a[href]",
                "div[role='button']",
                "button",
                "span[role='button']",
            ];
            for (const selector of actionSelectors) {
                for (const node of root.querySelectorAll(selector)) {
                    if (!isVisible(node)) continue;
                    const cleaned = [
                        node.getAttribute('aria-label') || '',
                        node.getAttribute('title') || '',
                        node.textContent || '',
                    ].join(' ').replace(/\\s+/g, ' ').trim();
                    if (!cleaned || cleaned.length > 120) continue;
                    if (!(/[0-9]/.test(cleaned) && /(comment|share|reaction|like|reply)/i.test(cleaned))) continue;
                    if (seenActionTexts.has(cleaned)) continue;
                    seenActionTexts.add(cleaned);
                    actionTexts.push(cleaned);
                }
            }

            const dateCandidates = [];
            const seenDates = new Set();
            const addDate = (value) => {
                const cleaned = (value || '').trim();
                if (!cleaned || seenDates.has(cleaned)) return;
                seenDates.add(cleaned);
                dateCandidates.push(cleaned);
            };
            for (const node of root.querySelectorAll("abbr, time, a[aria-label], span[aria-label]")) {
                addDate(node.getAttribute('aria-label'));
                addDate(node.getAttribute('datetime'));
                addDate(node.getAttribute('data-tooltip-content'));
                addDate(node.textContent);
            }

            return {
                found: true,
                scopeType: best.type,
                matchedTarget: best.matchedTarget,
                matchedSlug: best.matchedSlug,
                scopeText: rootText,
                metricTexts,
                actionTexts,
                dateCandidates,
                preview: rootText.slice(0, 240),
            };
        }
    """
    try:
        return page.evaluate(script, arg=target_url) or {}
    except Exception:
        return {
            "found": False,
            "scopeType": "",
            "matchedTarget": False,
            "matchedSlug": False,
            "scopeText": "",
            "metricTexts": [],
            "actionTexts": [],
            "dateCandidates": [],
            "preview": "",
        }


def extract_post_date_from_snapshot(scope_snapshot: dict[str, Any]) -> tuple[str, Optional[datetime]]:
    for candidate in scope_snapshot.get("dateCandidates") or []:
        parsed = parse_facebook_datetime(candidate)
        if parsed is not None:
            return candidate, parsed
    return "", None


def focus_target_post(page, target_url: str, log_hook: Optional[LogHook] = None) -> None:
    script = """
        (targetUrl) => {
            const normalizeHref = (value) => {
                if (!value) return '';
                try {
                    const url = new URL(value, window.location.origin);
                    for (const key of ['comment_id', 'reply_comment_id', '__tn__', '__cft__', 'ref', 'refsrc', 'notif_t', 'comment_tracking', 'acontext']) {
                        url.searchParams.delete(key);
                    }
                    return `${url.origin}${url.pathname}${url.search}`;
                } catch (error) {
                    return String(value || '').split('&__')[0];
                }
            };
            const isVisible = (node) => {
                if (!node) return false;
                const style = window.getComputedStyle(node);
                if (!style || style.visibility === 'hidden' || style.display === 'none') return false;
                const rect = node.getBoundingClientRect();
                return rect.width > 8 && rect.height > 8;
            };
            const target = normalizeHref(targetUrl);
            let targetPath = '';
            try {
                if (target) {
                    const parsed = new URL(target);
                    targetPath = `${parsed.pathname}${parsed.search}`;
                }
            } catch (error) {}
            const roots = Array.from(document.querySelectorAll("div[role='dialog'], div[role='main'] div[role='article'], div[role='feed'] div[role='article'], div[data-pagelet*='FeedUnit']"));
            for (const root of roots) {
                if (!isVisible(root)) continue;
                const links = Array.from(root.querySelectorAll("a[href]"))
                    .map((anchor) => normalizeHref(anchor.href))
                    .filter(Boolean);
                const matched = Boolean(target && links.some((href) => href === target || (targetPath && href.includes(targetPath))));
                if (!matched) continue;
                root.scrollIntoView({ block: 'center', inline: 'nearest' });
                return true;
            }
            return false;
        }
    """
    try:
        focused = bool(page.evaluate(script, arg=target_url))
        if focused:
            page.wait_for_timeout(300)
            emit_log(log_hook, "INFO", "Active post focused", "Scrolled the target Facebook post into view before extraction.")
    except Exception:
        return


def activate_target_post_surface(page, target_url: str, log_hook: Optional[LogHook] = None) -> bool:
    script = """
        (targetUrl) => {
            const normalizeHref = (value) => {
                if (!value) return '';
                try {
                    const url = new URL(value, window.location.origin);
                    for (const key of ['comment_id', 'reply_comment_id', '__tn__', '__cft__', 'ref', 'refsrc', 'notif_t', 'comment_tracking', 'acontext']) {
                        url.searchParams.delete(key);
                    }
                    return `${url.origin}${url.pathname}${url.search}`;
                } catch (error) {
                    return String(value || '').split('&__')[0];
                }
            };
            const isVisible = (node) => {
                if (!node) return false;
                const style = window.getComputedStyle(node);
                if (!style || style.visibility === 'hidden' || style.display === 'none') return false;
                const rect = node.getBoundingClientRect();
                return rect.width > 8 && rect.height > 8;
            };
            const target = normalizeHref(targetUrl);
            let targetPath = '';
            try {
                if (target) {
                    const parsed = new URL(target);
                    targetPath = `${parsed.pathname}${parsed.search}`;
                }
            } catch (error) {}
            const anchors = Array.from(document.querySelectorAll("a[href]")).filter((anchor) => {
                if (!isVisible(anchor)) return false;
                const href = normalizeHref(anchor.href);
                return Boolean(href && target && (href === target || (targetPath && href.includes(targetPath))));
            });
            const anchor = anchors[0];
            if (!anchor) return false;
            anchor.scrollIntoView({ block: 'center', inline: 'nearest' });
            anchor.click();
            return true;
        }
    """
    try:
        activated = bool(page.evaluate(script, arg=target_url))
        if activated:
            page.wait_for_timeout(500)
            emit_log(log_hook, "INFO", "Post modal detected", "Activated the matching Facebook post surface before extraction.")
        return activated
    except Exception:
        return False


def extract_post_date(page, target_url: str = "") -> tuple[str, Optional[datetime], dict[str, Any]]:
    snapshot = inspect_active_post_scope(page, target_url=target_url)
    raw_value, parsed = extract_post_date_from_snapshot(snapshot)
    return raw_value, parsed, snapshot


def select_all_comments_mode(page, target_url: str = "", log_hook: Optional[LogHook] = None) -> None:
    script = """
        (targetUrl) => {
            const normalizeHref = (value) => {
                if (!value) return '';
                try {
                    const url = new URL(value, window.location.origin);
                    for (const key of ['comment_id', 'reply_comment_id', '__tn__', '__cft__', 'ref', 'refsrc', 'notif_t', 'comment_tracking', 'acontext']) {
                        url.searchParams.delete(key);
                    }
                    return `${url.origin}${url.pathname}${url.search}`;
                } catch (error) {
                    return String(value || '').split('&__')[0];
                }
            };
            const isVisible = (node) => {
                if (!node) return false;
                const style = window.getComputedStyle(node);
                if (!style || style.visibility === 'hidden' || style.display === 'none') return false;
                const rect = node.getBoundingClientRect();
                return rect.width > 8 && rect.height > 8;
            };
            const target = normalizeHref(targetUrl);
            let targetPath = '';
            try {
                if (target) {
                    const parsed = new URL(target);
                    targetPath = `${parsed.pathname}${parsed.search}`;
                }
            } catch (error) {}
            const roots = Array.from(document.querySelectorAll("div[role='dialog'], div[role='main'] div[role='article'], div[role='feed'] div[role='article'], div[data-pagelet*='FeedUnit']"));
            let root = null;
            for (const candidate of roots) {
                if (!isVisible(candidate)) continue;
                const links = Array.from(candidate.querySelectorAll("a[href]"))
                    .map((anchor) => normalizeHref(anchor.href))
                    .filter(Boolean);
                if (target && links.some((href) => href === target || (targetPath && href.includes(targetPath)))) {
                    root = candidate;
                    break;
                }
            }
            if (!root) {
                root = roots.find(isVisible) || document.querySelector("div[role='main']") || document.body;
            }

            const collectText = (node) => (node.innerText || node.textContent || '').replace(/\\s+/g, ' ').trim().toLowerCase();
            const directOption = Array.from(document.querySelectorAll("[role='menuitem'], [role='option'], div[role='button'], button, span"))
                .find((node) => isVisible(node) && collectText(node).includes('all comments'));
            if (directOption) {
                directOption.click();
                return "All comments selected";
            }

            const sortControl = Array.from(root.querySelectorAll("div[role='button'], button, span[role='button'], a[role='button'], span"))
                .find((node) => {
                    if (!isVisible(node)) return false;
                    const text = collectText(node);
                    return text.includes('most relevant') || text.includes('top comments') || text.includes('newest') || text.includes('all comments');
                });
            if (sortControl) {
                const text = collectText(sortControl);
                if (text.includes('all comments')) {
                    return "All comments already selected";
                }
                (sortControl.closest("div[role='button'], button, a[role='button']") || sortControl).click();
                return "Opened comment sort menu";
            }

            return "";
        }
    """
    try:
        result = (page.evaluate(script, arg=target_url) or "").strip()
    except Exception:
        result = ""
    if not result:
        return
    page.wait_for_timeout(COMMENT_LOAD_WAIT_MS)
    emit_log(log_hook, "INFO", "All comments", result)


def count_visible_comment_nodes(page, target_url: str = "") -> int:
    script = """
        (targetUrl) => {
            const normalizeHref = (value) => {
                if (!value) return '';
                try {
                    const url = new URL(value, window.location.origin);
                    for (const key of ['comment_id', 'reply_comment_id', '__tn__', '__cft__', 'ref', 'refsrc', 'notif_t', 'comment_tracking', 'acontext']) {
                        url.searchParams.delete(key);
                    }
                    return `${url.origin}${url.pathname}${url.search}`;
                } catch (error) {
                    return String(value || '').split('&__')[0];
                }
            };
            const isVisible = (node) => {
                if (!node) return false;
                const style = window.getComputedStyle(node);
                if (!style || style.visibility === 'hidden' || style.display === 'none') return false;
                const rect = node.getBoundingClientRect();
                return rect.width > 8 && rect.height > 8;
            };
            const target = normalizeHref(targetUrl);
            let targetPath = '';
            try {
                if (target) {
                    const parsed = new URL(target);
                    targetPath = `${parsed.pathname}${parsed.search}`;
                }
            } catch (error) {}
            const roots = Array.from(document.querySelectorAll("div[role='dialog'], div[role='main'] div[role='article'], div[role='feed'] div[role='article'], div[data-pagelet*='FeedUnit']"));
            let root = null;
            for (const candidate of roots) {
                if (!isVisible(candidate)) continue;
                const links = Array.from(candidate.querySelectorAll("a[href]"))
                    .map((anchor) => normalizeHref(anchor.href))
                    .filter(Boolean);
                if (target && links.some((href) => href === target || (targetPath && href.includes(targetPath)))) {
                    root = candidate;
                    break;
                }
            }
            if (!root) {
                root = roots.find(isVisible) || document.querySelector("div[role='main']") || document.body;
            }

            const candidates = new Set();
            for (const selector of ["ul li", "div[aria-label*='Comment']", "div[role='article'] ul li"]) {
                for (const node of root.querySelectorAll(selector)) {
                    const text = (node.innerText || '').trim();
                    if (!text || text.length < 8) continue;
                    candidates.add(node);
                }
            }
            return candidates.size;
        }
    """
    try:
        return int(page.evaluate(script, arg=target_url) or 0)
    except Exception:
        return 0


def expand_visible_comment_threads(
    page,
    target_url: str = "",
    log_hook: Optional[LogHook] = None,
    rounds: int = COMMENT_EXPANSION_ROUNDS,
) -> None:
    select_all_comments_mode(page, target_url=target_url, log_hook=log_hook)

    click_script = """
        (targetUrl) => {
            const normalizeHref = (value) => {
                if (!value) return '';
                try {
                    const url = new URL(value, window.location.origin);
                    for (const key of ['comment_id', 'reply_comment_id', '__tn__', '__cft__', 'ref', 'refsrc', 'notif_t', 'comment_tracking', 'acontext']) {
                        url.searchParams.delete(key);
                    }
                    return `${url.origin}${url.pathname}${url.search}`;
                } catch (error) {
                    return String(value || '').split('&__')[0];
                }
            };
            const isVisible = (node) => {
                if (!node) return false;
                const style = window.getComputedStyle(node);
                if (!style || style.visibility === 'hidden' || style.display === 'none') return false;
                const rect = node.getBoundingClientRect();
                return rect.width > 8 && rect.height > 8;
            };
            const target = normalizeHref(targetUrl);
            let targetPath = '';
            try {
                if (target) {
                    const parsed = new URL(target);
                    targetPath = `${parsed.pathname}${parsed.search}`;
                }
            } catch (error) {}
            const roots = Array.from(document.querySelectorAll("div[role='dialog'], div[role='main'] div[role='article'], div[role='feed'] div[role='article'], div[data-pagelet*='FeedUnit']"));
            let root = null;
            for (const candidate of roots) {
                if (!isVisible(candidate)) continue;
                const links = Array.from(candidate.querySelectorAll("a[href]"))
                    .map((anchor) => normalizeHref(anchor.href))
                    .filter(Boolean);
                if (target && links.some((href) => href === target || (targetPath && href.includes(targetPath)))) {
                    root = candidate;
                    break;
                }
            }
            if (!root) {
                root = roots.find(isVisible) || document.querySelector("div[role='main']") || document.body;
            }

            const actions = [
                { label: 'View more comments', match: (text) => text.includes('view more comments') || text.includes('see more comments') || text.includes('view previous comments') || text.includes('more comments') },
                { label: 'View replies', match: (text) => text.includes('view replies') || text.includes('view more replies') || text.includes('see more replies') || text.includes('view previous replies') || text.includes('more replies') },
                { label: 'See more', match: (text, node) => (text === 'see more' || text.startsWith('see more ')) && Boolean(node.closest("ul li, div[aria-label*='Comment'], div[role='article'] ul li")) },
            ];

            const nodes = Array.from(root.querySelectorAll("div[role='button'], button, span[role='button'], a[role='button'], span"));
            for (const action of actions) {
                for (const node of nodes) {
                    if (!isVisible(node)) continue;
                    const text = (node.innerText || node.textContent || '').replace(/\\s+/g, ' ').trim().toLowerCase();
                    if (!text || text.length > 80) continue;
                    if (!action.match(text, node)) continue;
                    const targetNode = node.closest("div[role='button'], button, a[role='button']") || node;
                    targetNode.scrollIntoView({ block: 'center', inline: 'nearest' });
                    targetNode.click();
                    return action.label;
                }
            }
            return '';
        }
    """

    previous_count = count_visible_comment_nodes(page, target_url=target_url)
    no_growth_rounds = 0
    for round_index in range(1, rounds + 1):
        if round_index > 1:
            try:
                page.mouse.wheel(0, 1100)
            except Exception:
                page.evaluate("window.scrollBy(0, 1100);")

        try:
            clicked_label = (page.evaluate(click_script, arg=target_url) or "").strip()
        except Exception:
            clicked_label = ""

        page.wait_for_timeout(COMMENT_LOAD_WAIT_MS + 300)
        current_count = count_visible_comment_nodes(page, target_url=target_url)
        if clicked_label:
            emit_log(
                log_hook,
                "INFO",
                "Comments expanded",
                f"Round {round_index}: clicked {clicked_label} (visible nodes={current_count}).",
            )
            no_growth_rounds = 0
        elif current_count > previous_count:
            emit_log(
                log_hook,
                "INFO",
                "Comments expanded",
                f"Round {round_index}: more visible comment nodes loaded (visible nodes={current_count}).",
            )
            no_growth_rounds = 0
        else:
            emit_log(log_hook, "INFO", "Comments expanded", f"Round {round_index}: no additional visible comments or replies.")
            no_growth_rounds += 1

        if no_growth_rounds >= 2:
            emit_log(log_hook, "INFO", "Comments expanded", "Stopping comment expansion after repeated no-growth rounds.")
            break

        previous_count = current_count


def extract_visible_comments(page, post_url: str, limit: int = 50, log_hook: Optional[LogHook] = None) -> list[CommentData]:
    expand_visible_comment_threads(page, target_url=post_url, log_hook=log_hook)
    script = """
        (payload) => {
            const { postUrl, limit } = payload;
            const normalizeHref = (value) => {
                if (!value) return '';
                try {
                    const url = new URL(value, window.location.origin);
                    for (const key of ['comment_id', 'reply_comment_id', '__tn__', '__cft__', 'ref', 'refsrc', 'notif_t', 'comment_tracking', 'acontext']) {
                        url.searchParams.delete(key);
                    }
                    return `${url.origin}${url.pathname}${url.search}`;
                } catch (error) {
                    return String(value || '').split('&__')[0];
                }
            };
            const isVisible = (node) => {
                if (!node) return false;
                const style = window.getComputedStyle(node);
                if (!style || style.visibility === 'hidden' || style.display === 'none') return false;
                const rect = node.getBoundingClientRect();
                return rect.width > 8 && rect.height > 8;
            };
            const target = normalizeHref(postUrl);
            let targetPath = '';
            try {
                if (target) {
                    const parsed = new URL(target);
                    targetPath = `${parsed.pathname}${parsed.search}`;
                }
            } catch (error) {}
            const roots = Array.from(document.querySelectorAll("div[role='dialog'], div[role='main'] div[role='article'], div[role='feed'] div[role='article'], div[data-pagelet*='FeedUnit']"));
            let root = null;
            for (const candidate of roots) {
                if (!isVisible(candidate)) continue;
                const links = Array.from(candidate.querySelectorAll("a[href]"))
                    .map((anchor) => normalizeHref(anchor.href))
                    .filter(Boolean);
                if (target && links.some((href) => href === target || (targetPath && href.includes(targetPath)))) {
                    root = candidate;
                    break;
                }
            }
            if (!root) {
                root = roots.find(isVisible) || document.querySelector("div[role='main']") || document.body;
            }

            const results = [];
            const seen = new Set();
            const datePattern = /(just now|\\d+\\s*(?:s|m|h|d|w)|\\d+\\s+(?:second|minute|hour|day|week|month|year)s?\\s+ago|yesterday|today|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)/i;
            const uiTokens = new Set(['like', 'reply', 'author', 'edited', 'top fan', 'follow', 'message']);
            const candidates = [];
            const commentRoots = Array.from(root.querySelectorAll("div[aria-label*='Comment'], ul, div[role='article'] ul"))
                .filter((node) => isVisible(node) && /(comment|reply)/i.test((node.innerText || '').slice(0, 250)));
            const searchRoots = commentRoots.length ? commentRoots : [root];
            for (const searchRoot of searchRoots) {
                for (const selector of ["ul li", "div[aria-label*='Comment']"]) {
                    for (const node of searchRoot.querySelectorAll(selector)) {
                        if (isVisible(node)) candidates.push(node);
                    }
                }
            }

            for (const node of candidates) {
                const text = (node.innerText || '').trim();
                if (!text || text.length < 8 || text.length > 800) continue;
                const lines = text.split('\\n').map((line) => line.trim()).filter(Boolean);
                if (!lines.length) continue;

                let commenter = '';
                const authorCandidate = node.querySelector("strong a, h3 a, a[role='link']");
                if (authorCandidate) {
                    commenter = (authorCandidate.textContent || '').trim();
                }
                if (!commenter) {
                    commenter = lines[0] || '';
                }

                let commentDate = '';
                const dateNode = node.querySelector("abbr, time, a[aria-label], span[aria-label]");
                if (dateNode) {
                    commentDate = (
                        dateNode.getAttribute('aria-label') ||
                        dateNode.getAttribute('datetime') ||
                        dateNode.textContent ||
                        ''
                    ).trim();
                }
                if (!commentDate) {
                    const lineMatch = lines.find((line) => datePattern.test(line));
                    commentDate = lineMatch || '';
                }

                const threadType = node.closest('ul ul') ? 'Reply' : 'Comment';
                const filtered = lines.filter((line, index) => {
                    const lower = line.toLowerCase();
                    if (!line) return false;
                    if (index === 0 && line === commenter) return false;
                    if (commentDate && line === commentDate) return false;
                    if (uiTokens.has(lower)) return false;
                    if (lower.startsWith('view more repl') || lower.startsWith('see more repl')) return false;
                    if (lower.startsWith('view more comment') || lower.startsWith('see more comment')) return false;
                    if (lower === 'see more') return false;
                    if (lower.startsWith('facebook facebook facebook')) return false;
                    return true;
                });

                const commentText = filtered.join(' ').trim() || text;
                const identity = `${threadType}|${commenter}|${commentDate}|${commentText}`;
                if (!commentText || seen.has(identity)) continue;
                seen.add(identity);
                results.push({ commenter, commentText, commentDate, threadType });
                if (results.length >= limit) break;
            }
            return results;
        }
    """
    try:
        rows = page.evaluate(script, arg={"postUrl": post_url, "limit": limit})
    except Exception:
        return []

    comments: list[CommentData] = []
    for row in rows or []:
        comment_text = (row.get("commentText") or "").strip()
        commenter = (row.get("commenter") or "").strip()
        if not comment_text:
            continue
        comments.append(
            CommentData(
                post_url=post_url,
                commenter_name=commenter or "Unknown",
                comment_text=comment_text,
                comment_date_raw=(row.get("commentDate") or "").strip(),
                thread_type=(row.get("threadType") or "Comment").strip() or "Comment",
            )
        )
    return comments


def extract_metric_from_texts(texts: list[str], patterns: list[str]) -> Optional[int]:
    seen: set[str] = set()
    for raw_text in texts:
        cleaned = re.sub(r"\s+", " ", raw_text or "").strip()
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        for pattern in patterns:
            match = re.search(pattern, cleaned, re.IGNORECASE)
            if match:
                return parse_count(match.group(1))
    return None


def extract_text_metrics(page, target_url: str = "", scope_snapshot: Optional[dict[str, Any]] = None) -> tuple[Optional[int], Optional[int], Optional[int]]:
    snapshot = scope_snapshot or inspect_active_post_scope(page, target_url=target_url)
    metric_texts = [re.sub(r"\s+", " ", text or "").strip() for text in (snapshot.get("metricTexts") or []) if text]
    action_texts = [re.sub(r"\s+", " ", text or "").strip() for text in (snapshot.get("actionTexts") or []) if text]
    root_text = re.sub(r"\s+", " ", snapshot.get("scopeText") or "").strip()
    search_texts = action_texts + metric_texts + ([root_text] if root_text else [])

    reactions = extract_metric_from_texts(
        search_texts,
        [
            r"([\d.,KMBkmb]+)\s+reactions?",
            r"([\d.,KMBkmb]+)\s+people reacted",
            r"([\d.,KMBkmb]+)\s+likes?",
            r"reactions?\s+([\d.,KMBkmb]+)",
        ],
    )
    comments = extract_metric_from_texts(
        search_texts,
        [
            r"([\d.,KMBkmb]+)\s+comments?",
            r"([\d.,KMBkmb]+)\s+comment[s]?\b",
            r"comments?\s+([\d.,KMBkmb]+)",
        ],
    )
    shares = extract_metric_from_texts(
        search_texts,
        [
            r"([\d.,KMBkmb]+)\s+shares?",
            r"([\d.,KMBkmb]+)\s+share[s]?\b",
            r"shares?\s+([\d.,KMBkmb]+)",
        ],
    )

    return reactions, comments, shares


def wait_for_active_post_scope(page, target_url: str, timeout_ms: int = 5000) -> dict[str, Any]:
    deadline = time.perf_counter() + max(timeout_ms, 500) / 1000
    best_snapshot: dict[str, Any] = {}
    while time.perf_counter() < deadline:
        snapshot = inspect_active_post_scope(page, target_url=target_url)
        if snapshot.get("found"):
            best_snapshot = snapshot
            if (
                snapshot.get("matchedTarget")
                or snapshot.get("matchedSlug")
                or snapshot.get("dateCandidates")
                or snapshot.get("metricTexts")
                or snapshot.get("actionTexts")
            ):
                return snapshot
        page.wait_for_timeout(250)
    return best_snapshot


def open_post_for_extraction(
    page,
    url: str,
    goto_timeout: int = POST_GOTO_TIMEOUT,
    log_hook: Optional[LogHook] = None,
) -> tuple[str, Optional[datetime], str, dict[str, Any]]:
    page.goto(url, wait_until="domcontentloaded", timeout=goto_timeout)
    try:
        page.wait_for_load_state("networkidle", timeout=12_000)
    except Exception:
        pass
    page.wait_for_timeout(1200)
    wait_for_selector(page, "body", 2000)
    apply_local_page_preferences(page)
    page.wait_for_timeout(450)
    checkpoint_required, checkpoint_reason = detect_checkpoint_or_verification(page)
    if checkpoint_required:
        emit_log(log_hook, "WARN", "Facebook checkpoint", checkpoint_reason)
        raise AuthRequiredError("waiting_verification", checkpoint_reason)
    login_required, login_reason = detect_login_gate(page)
    if login_required:
        emit_log(log_hook, "WARN", "Facebook login required", login_reason)
        raise AuthRequiredError("waiting_login", login_reason)
    focus_target_post(page, url, log_hook=log_hook)
    page.wait_for_timeout(600)
    post_type = infer_post_type(url)
    scope_snapshot = wait_for_active_post_scope(page, url)
    if scope_snapshot.get("found") and not (
        scope_snapshot.get("matchedTarget")
        or scope_snapshot.get("matchedSlug")
        or scope_snapshot.get("metricTexts")
        or scope_snapshot.get("actionTexts")
    ):
        if activate_target_post_surface(page, url, log_hook=log_hook):
            scope_snapshot = wait_for_active_post_scope(page, url, timeout_ms=6000)
    raw_date, date_obj = extract_post_date_from_snapshot(scope_snapshot)
    if scope_snapshot.get("found"):
        emit_log(
            log_hook,
            "INFO",
            "Post surface detected",
            f"Using {scope_snapshot.get('scopeType', 'post')} scoped extraction for the active Facebook post.",
        )
    return raw_date, date_obj, post_type, scope_snapshot


def extract_metrics_from_loaded_post(
    page,
    url: str,
    raw_date: str,
    date_obj: Optional[datetime],
    post_type: str,
    collection_type: str,
    log_hook: Optional[LogHook] = None,
    scope_snapshot: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    """
    Hardened per-post extraction with retries and DOM stabilization.
    
    Steps:
    1. Scroll post into view and center
    2. Wait for DOM to stabilize
    3. Retry extraction up to 3 times for each metric
    4. Only mark as Unavailable after all retries exhausted
    """
    max_retries = 3
    reactions = None
    comments_count = None
    shares = None
    unavailable_metrics = 0
    notes: list[str] = []
    
    # Step 1: Scroll post into view and center
    try:
        page.evaluate("""() => {
            const activePost = document.querySelector('[role="article"]') || 
                              document.querySelector('div[data-content-id]') ||
                              document.querySelector('div[role="feed"] > div > div');
            if (activePost) {
                activePost.scrollIntoView({block: 'center', behavior: 'smooth'});
            }
        }""")
        page.wait_for_timeout(300)
    except Exception:
        pass
    
    # Step 2: Wait for DOM stabilization
    wait_for_scroll_stabilization(page, timeout_ms=3000)
    page.wait_for_timeout(400)
    
    # Step 3: Retry extraction for each metric
    for attempt in range(1, max_retries + 1):
        emit_log(log_hook, "INFO", f"Extraction attempt {attempt}/{max_retries}", f"Extracting metrics for {url[:60]}...")
        
        snapshot = scope_snapshot or inspect_active_post_scope(page, target_url=url)
        reactions, comments_count, shares = extract_text_metrics(page, target_url=url, scope_snapshot=snapshot)
        
        # Check if we got all metrics
        metrics_found = 0
        if reactions is not None:
            metrics_found += 1
        if comments_count is not None:
            metrics_found += 1
        if shares is not None:
            metrics_found += 1
        
        # If all metrics found, stop retrying
        if metrics_found == 3:
            emit_log(log_hook, "SUCCESS", "Metrics extracted", f"All metrics found on attempt {attempt}")
            break
        
        # If this isn't the last attempt, wait before retrying
        if attempt < max_retries:
            wait_ms = 300 + (attempt * 150)  # Increasing wait: 450ms, 600ms, 750ms
            page.wait_for_timeout(wait_ms)
            # Scroll slightly to trigger re-render
            page.evaluate("() => window.scrollBy(0, 10)")
            page.wait_for_timeout(200)
        else:
            emit_log(log_hook, "WARN", "Metrics incomplete", f"After {max_retries} attempts: reactions={reactions}, comments={comments_count}, shares={shares}")
    
    # Step 4: Mark unavailable only after all retries
    if reactions is None:
        notes.append("Reactions unavailable (not visible after 3 retries)")
        emit_log(log_hook, "WARN", "Metric unavailable", f"Reactions | post_url={url[:60]} | reason=not visible after retries")
        unavailable_metrics += 1
    else:
        emit_log(log_hook, "INFO", "Reactions extracted", str(reactions))
        
    if comments_count is None:
        notes.append("Comments count unavailable (not visible after 3 retries)")
        emit_log(log_hook, "WARN", "Metric unavailable", f"Comments | post_url={url[:60]} | reason=not visible after retries")
        unavailable_metrics += 1
    else:
        emit_log(log_hook, "INFO", "Comments count extracted", str(comments_count))
    if shares is None:
        notes.append("Shares unavailable (not visible after 3 retries)")
        emit_log(log_hook, "WARN", "Metric unavailable", f"Shares | post_url={url[:60]} | reason=not visible after retries")
        unavailable_metrics += 1
    else:
        emit_log(log_hook, "INFO", "Shares extracted", str(shares))

    comments_preview: list[CommentData] = []
    # Facebook now collects posts_only - no comment collection
    
    emit_log(
        log_hook,
        "INFO",
        "Metrics extracted",
        (
            f"{url} -> reactions={reactions if reactions is not None else 'Unavailable'}, "
            f"comments={comments_count if comments_count is not None else 'Unavailable'}, "
            f"shares={shares if shares is not None else 'Unavailable'}, "
            f"date={raw_date or 'Unavailable'}"
        ),
    )

    return {
        "post_link": url,
        "post_date": raw_date or "N/A",
        "post_date_obj": date_obj,
        "post_type": post_type,
        "reactions": reactions if reactions is not None else "N/A",
        "comments_count": comments_count if comments_count is not None else "N/A",
        "shares": shares if shares is not None else "N/A",
        "notes": notes,
        "unavailable_metrics": unavailable_metrics,
        "comments_preview": comments_preview,
    }


def extract_post_from_feed(
    page,
    url: str,
    collection_type: str,
    log_hook: Optional[LogHook] = None,
) -> Optional[dict[str, Any]]:
    snapshot = inspect_active_post_scope(page, target_url=url)
    if not snapshot.get("found"):
        return None
    if not (snapshot.get("matchedTarget") or snapshot.get("matchedSlug")):
        return None

    raw_date, date_obj = extract_post_date_from_snapshot(snapshot)
    post_type = infer_post_type(url)
    return extract_metrics_from_loaded_post(
        page,
        url,
        raw_date,
        date_obj,
        post_type,
        collection_type,
        log_hook=log_hook,
        scope_snapshot=snapshot,
    )


def classify_post_date_coverage(
    post_date_obj: Optional[datetime],
    start_date: datetime,
    end_date: Optional[datetime],
) -> tuple[str, str]:
    if post_date_obj is None:
        return "unknown_date", "Skipped because the Facebook post date could not be detected."
    if end_date is not None and post_date_obj > end_date:
        return "newer_than_end", f"Skipped because {post_date_obj.strftime(DATE_INPUT_FORMAT)} is newer than end date {end_date.strftime(DATE_INPUT_FORMAT)}."
    if post_date_obj < start_date:
        return "older_than_start", f"Skipped because {post_date_obj.strftime(DATE_INPUT_FORMAT)} is older than start date {start_date.strftime(DATE_INPUT_FORMAT)}."
    return "in_range", f"Included because {post_date_obj.strftime(DATE_INPUT_FORMAT)} is within the selected Facebook date coverage."


def post_matches_date_coverage(post: PostData, start_date: datetime, end_date: Optional[datetime]) -> bool:
    return classify_post_date_coverage(post.post_date_obj, start_date, end_date)[0] == "in_range"


def format_post_date(post: PostData) -> str:
    if post.post_date_obj is not None:
        return post.post_date_obj.strftime(DATE_INPUT_FORMAT)
    return post.post_date_raw or "Cannot detect"


LAST_RUN_DIAGNOSTICS: dict[str, Any] = {}


def set_run_diagnostics(payload: dict[str, Any]) -> None:
    LAST_RUN_DIAGNOSTICS.clear()
    LAST_RUN_DIAGNOSTICS.update(payload)


def sanitize_excel_value(value: Any, field_name: str = "") -> Any:
    """
    Sanitize a value for Excel export. Converts non-scalar values to JSON strings.
    
    Args:
        value: The value to sanitize
        field_name: The field name for logging purposes
        
    Returns:
        A scalar value (str, int, float, bool, None) safe for Excel
    """
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        return value
    
    # Convert dict/list/object to JSON string
    import json
    try:
        if isinstance(value, (dict, list)):
            json_str = json.dumps(value)
            if field_name:
                import sys
                print(f"[SANITIZE] Field '{field_name}' converted {type(value).__name__} to JSON", file=sys.stderr)
            return json_str
    except Exception:
        pass
    
    # Fallback: convert to string
    if field_name:
        import sys
        print(f"[SANITIZE] Field '{field_name}' converted {type(value).__name__} to string", file=sys.stderr)
    return str(value)


def sanitize_facebook_dataset(posts: list[Any]) -> list[Any]:
    """
    GLOBAL DATA SANITIZATION: Recursively sanitize entire dataset before export.
    
    Converts all dict/list values to safe types for Excel.
    Ensures no non-scalar values reach openpyxl.
    """
    sanitized_posts = []
    
    for post in posts:
        if isinstance(post, dict):
            sanitized_post = {}
            for key, value in post.items():
                # Sanitize each field
                if key in ["post_link", "post_date", "post_type", "reactions", "comments_count", "shares", "notes"]:
                    safe_value = sanitize_excel_value(value, key)
                    sanitized_post[key] = safe_value
                    if str(type(value)) != str(type(safe_value)):
                        # Log that we sanitized this field
                        pass  # Logging happens in sanitize_excel_value
                else:
                    # Unknown field - apply sanitization anyway
                    sanitized_post[key] = sanitize_excel_value(value, key)
            sanitized_posts.append(sanitized_post)
        else:
            # Non-dict post object - sanitize field by field
            sanitized_post = {
                "post_link": sanitize_excel_value(getattr(post, "post_link", getattr(post, "url", "")), "post_link"),
                "post_date": sanitize_excel_value(getattr(post, "post_date", getattr(post, "date", "")), "post_date"),
                "post_type": sanitize_excel_value(getattr(post, "post_type", ""), "post_type"),
                "reactions": sanitize_excel_value(getattr(post, "reactions", "N/A"), "reactions"),
                "comments_count": sanitize_excel_value(getattr(post, "comments_count", "N/A"), "comments_count"),
                "shares": sanitize_excel_value(getattr(post, "shares", "N/A"), "shares"),
                "notes": sanitize_excel_value(getattr(post, "notes", ""), "notes"),
            }
            sanitized_posts.append(sanitized_post)
    
    return sanitized_posts


def save_facebook_excel(posts: list[Any], filename: str, coverage_label: str, collection_type: str) -> None:
    # GLOBAL SANITIZATION: Apply to entire dataset before writing
    posts = sanitize_facebook_dataset(posts)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Facebook Posts"

    ws["A1"] = coverage_label
    ws["A2"] = f"Collection type: {collection_type.replace('_', ' ')}"
    ws.append([])
    ws.append(["Post Link", "Post Date", "Post Type", "Reactions", "Comments Count", "Shares", "Notes"])

    for post in posts:
        if isinstance(post, dict):
            ws.append([
                post.get("post_link", ""),
                post.get("post_date", "N/A"),
                post.get("post_type", ""),
                post.get("reactions", "N/A"),
                post.get("comments_count", "N/A"),
                post.get("shares", "N/A"),
                post.get("notes", ""),
            ])

    ws.column_dimensions["A"].width = 64
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 44

    # Comments sheet removed - Facebook now collects posts_only
    
    diagnostics_sheet = wb.create_sheet("Diagnostics")
    diagnostics_sheet.append(["Metric", "Value"])
    for key, value in LAST_RUN_DIAGNOSTICS.items():
        diagnostics_sheet.append([key, value])
    diagnostics_sheet.column_dimensions["A"].width = 32
    diagnostics_sheet.column_dimensions["B"].width = 64

    wb.save(filename)


def save_empty_result_excel(
    filename: str,
    coverage_label: str,
    total_links_collected: int,
    oldest_detected: Optional[datetime],
    newest_detected: Optional[datetime],
    reason: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Facebook Diagnostics"

    ws["A1"] = "No posts found within selected date range."
    ws["A2"] = coverage_label
    ws["A4"] = "Reason"
    ws["B4"] = reason
    ws["A5"] = "Total links collected"
    ws["B5"] = total_links_collected
    ws["A6"] = "Newest detected post date"
    ws["B6"] = newest_detected.strftime(DATE_INPUT_FORMAT) if newest_detected else "Unknown"
    ws["A7"] = "Oldest detected post date"
    ws["B7"] = oldest_detected.strftime(DATE_INPUT_FORMAT) if oldest_detected else "Unknown"

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 70
    wb.save(filename)
