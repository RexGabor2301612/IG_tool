from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from collections import defaultdict
from pathlib import Path
from typing import Any, Callable, List, Optional
from urllib.parse import urlparse

from openpyxl import Workbook
from playwright.sync_api import Error as PlaywrightError, sync_playwright


PROFILE_URL = "https://www.instagram.com/cebuanalhuillier/"
OUTPUT_FILE = "instagram_grouped_by_month.xlsx"
# Persistent browser profiles are intentionally not used in cloud hosting.
# Use PLAYWRIGHT_STORAGE_STATE=/path/to/state.json only if you intentionally
# provide an exported login state file in your deployment environment.
PLAYWRIGHT_STORAGE_STATE = os.getenv("PLAYWRIGHT_STORAGE_STATE", "").strip() or None
PLAYWRIGHT_HEADLESS = os.getenv("PLAYWRIGHT_HEADLESS", "true").strip().lower() not in {"0", "false", "no", "off"}
PLAYWRIGHT_AUTO_INSTALL = os.getenv("PLAYWRIGHT_AUTO_INSTALL", "true").strip().lower() not in {"0", "false", "no", "off"}

# -----------------------------------------------------------------------------
# Instagram login credentials
# Prefer environment variables in production / cloud deployments:
#   INSTAGRAM_USERNAME
#   INSTAGRAM_PASSWORD
#
# PUT TEST ACCOUNT LOGIN DETAILS HERE only if you intentionally need a local
# development fallback and you understand the security tradeoff. Keep them
# backend-only and never expose them to HTML, CSS, or JavaScript.
#
# Example local-only fallback:
# TEST_INSTAGRAM_USERNAME = ""
# TEST_INSTAGRAM_PASSWORD = ""
# -----------------------------------------------------------------------------
TEST_INSTAGRAM_USERNAME = ""
TEST_INSTAGRAM_PASSWORD = ""
INSTAGRAM_USERNAME = os.getenv("INSTAGRAM_USERNAME", "").strip() or TEST_INSTAGRAM_USERNAME.strip()
INSTAGRAM_PASSWORD = os.getenv("INSTAGRAM_PASSWORD", "").strip() or TEST_INSTAGRAM_PASSWORD.strip()
# Only collect posts from this date onwards (Instagram shows newest first, so older posts appear later in scroll).
START_DATE = datetime(2026, 1, 1)
# Set to None to collect every discoverable post link during the crawl window.
MAX_POSTS = None
# Hard stops prevent endless scrolling on layout/load anomalies.
MAX_SCROLL_ROUNDS = 26
MAX_STAGNANT_ROUNDS = 5
# Stop scrolling after seeing this many old posts in a row (since newest posts appear first).
OLD_POSTS_THRESHOLD = 3
# Per-post retries handle transient navigation/DOM timing failures.
POST_LOAD_RETRIES = 3
BASE_POST_DELAY = 0.15
MAX_POST_DELAY = 0.6
PROFILE_LINK_WAIT_TIMEOUT = 9000
PROFILE_SETTLE_MS = 300
PROFILE_RETRY_MS = 700
POST_GOTO_TIMEOUT = 30000
POST_TIME_WAIT_TIMEOUT = 3500
POST_ARTICLE_WAIT_TIMEOUT = 2200
POST_SETTLE_MS = 120
METRIC_READY_TIMEOUT = 900
METRIC_SETTLE_MS = 120
METRIC_FALLBACK_MS = 200
SCROLL_WAIT_TIMEOUT = 1800
SCROLL_FALLBACK_MS = 450
SCOPED_TEXT_TIMEOUT = 1600
BODY_TEXT_TIMEOUT = 2500
NEXT_DATA_TIMEOUT = 1800
BLOCKED_RESOURCE_TYPES = {"image", "media", "font"}
PROFILE_GRID_SELECTOR = "a[href*='/p/'], a[href*='/reel/']"
LOGIN_FORM_SELECTOR = "input[name='username'], input[name='password'], form[action*='/accounts/login/']"
POST_PRIMARY_SELECTOR = "time, article"
METRIC_SELECTOR = "svg[aria-label*='ike'], svg[aria-label*='omment'], article [role='button'] svg"
DATE_INPUT_FORMAT = "%Y-%m-%d"
VALID_INSTAGRAM_HOSTS = {"instagram.com", "www.instagram.com", "m.instagram.com"}
INVALID_FILENAME_CHARS = set('<>:"/\\|?*')
RESERVED_INSTAGRAM_PATHS = {
    "about",
    "accounts",
    "direct",
    "explore",
    "p",
    "privacy",
    "reel",
    "stories",
    "terms",
    "tv",
}
SLOW_SCROLL_SECONDS = 2.0
SLOW_POST_SECONDS = 4.0
LOGIN_FORM_TIMEOUT = 12000
LOGIN_POST_SUBMIT_TIMEOUT = 15000
LogHook = Callable[[str, str, str], None]
ProgressHook = Callable[[int, int, int], None]
LiveHook = Callable[[Any, str, dict[str, Any]], None]


def getenv_int(name: str, default: int) -> int:
    raw_value = os.getenv(name, "").strip()
    if not raw_value:
        return default

    try:
        return int(raw_value)
    except ValueError:
        return default


SCROLL_STAGNANT_LIMIT_OVERRIDE = max(0, getenv_int("SCROLL_STAGNANT_LIMIT", 0))
SCROLL_STOP_PROOF_ROUNDS = max(2, getenv_int("SCROLL_STOP_PROOF_ROUNDS", 2))
SCROLL_ATTEMPTS_PER_ROUND = max(3, getenv_int("SCROLL_ATTEMPTS_PER_ROUND", 5))
SCROLL_EXHAUSTION_CONFIRM_TIMEOUT = max(1500, getenv_int("SCROLL_EXHAUSTION_CONFIRM_TIMEOUT", 3500))
SCROLL_STRATEGIES = ("anchor-step", "viewport-step", "mouse-wheel", "page-down", "bottom-jump")


@dataclass
class PostData:
    url: str
    post_type: str
    post_date_raw: str
    post_date_obj: Optional[datetime]
    likes: Optional[int]
    comments: Optional[int]
    shares: int = 0  # Default to 0 if shares cannot be detected


@dataclass
class ScrapeConfig:
    profile_url: str
    scroll_rounds: int
    start_date: datetime
    end_date: Optional[datetime]
    output_file: str


def emit_log(log_hook: Optional[LogHook], level: str, action: str, details: str = "") -> None:
    if log_hook is None:
        return

    try:
        log_hook(level, action, details)
    except Exception:
        pass


def emit_progress(progress_hook: Optional[ProgressHook], scroll_round: int, total_rounds: int, posts_found: int) -> None:
    if progress_hook is None:
        return

    try:
        progress_hook(scroll_round, total_rounds, posts_found)
    except Exception:
        pass


def get_storage_state_path(require_exists: bool = False) -> Optional[Path]:
    if not PLAYWRIGHT_STORAGE_STATE:
        return None

    path = Path(PLAYWRIGHT_STORAGE_STATE)
    if require_exists and not path.exists():
        return None

    return path


def has_login_credentials() -> bool:
    return bool(INSTAGRAM_USERNAME and INSTAGRAM_PASSWORD)


def wait_for_selector(page, selector: str, timeout_ms: int) -> bool:
    try:
        page.locator(selector).first.wait_for(timeout=timeout_ms)
        return True
    except Exception:
        return False


def profile_grid_visible(page, timeout_ms: int = 400) -> bool:
    return wait_for_selector(page, PROFILE_GRID_SELECTOR, timeout_ms)


def detect_login_gate(page) -> tuple[bool, str]:
    if wait_for_selector(page, LOGIN_FORM_SELECTOR, 300):
        return True, "Instagram login form detected."

    gate_selectors = [
        ("button:has-text('Log in')", "Instagram login button detected."),
        ("button:has-text('Continue watching')", "Instagram continue-watching gate detected."),
        ("button:has-text('Sign up')", "Instagram sign-up gate detected."),
        ("text='Continue watching'", "Instagram continue-watching gate text detected."),
        ("text='Log in to continue'", "Instagram login wall detected."),
        ("text='See photos and videos'", "Instagram sign-in wall detected."),
    ]
    for selector, reason in gate_selectors:
        try:
            if page.locator(selector).first.count() > 0:
                return True, reason
        except Exception:
            continue

    try:
        body_text = page.locator("body").inner_text(timeout=500).lower()
    except Exception:
        body_text = ""

    if "continue watching" in body_text:
        return True, "Instagram continue-watching gate detected."
    if "log in to continue" in body_text:
        return True, "Instagram login wall detected."
    if "sign up to see photos and videos" in body_text:
        return True, "Instagram sign-up wall detected."

    return False, ""


def profile_ready_for_collection(page) -> bool:
    login_required, _ = detect_login_gate(page)
    if login_required:
        return False
    return profile_grid_visible(page, 350)


def collect_visible_post_links(page) -> List[str]:
    return page.evaluate(
        """() => {
            const seen = new Set();
            const results = [];
            for (const anchor of document.querySelectorAll("a[href*='/p/'], a[href*='/reel/']")) {
                const href = (anchor.href || "").split("?")[0];
                if (!href || seen.has(href)) continue;
                seen.add(href);
                results.push(href);
            }
            return results;
        }"""
    )


def wait_for_more_profile_links(page, previous_count: int, timeout_ms: int = SCROLL_WAIT_TIMEOUT) -> bool:
    try:
        page.wait_for_function(
            """(prev) => {
                const unique = new Set(
                    Array.from(document.querySelectorAll("a[href*='/p/'], a[href*='/reel/']"))
                        .map(anchor => (anchor.href || "").split("?")[0])
                        .filter(Boolean)
                );
                return unique.size > prev;
            }""",
            arg=previous_count,
            timeout=timeout_ms,
        )
        return True
    except Exception:
        return False


def get_effective_stagnant_limit(scroll_rounds: int) -> int:
    if SCROLL_STAGNANT_LIMIT_OVERRIDE > 0:
        return SCROLL_STAGNANT_LIMIT_OVERRIDE

    if scroll_rounds <= 0:
        return MAX_STAGNANT_ROUNDS

    return max(MAX_STAGNANT_ROUNDS, min(18, max(8, scroll_rounds // 3)))


def get_minimum_rounds_before_stop(scroll_rounds: int, stagnant_limit: int) -> int:
    if scroll_rounds <= 0:
        return stagnant_limit

    return min(scroll_rounds, max(8, min(15, stagnant_limit)))


def get_profile_scroll_state(page) -> dict[str, int | bool | str]:
    return page.evaluate(
        """() => {
            const anchors = Array.from(document.querySelectorAll("a[href*='/p/'], a[href*='/reel/']"));
            const fallback = document.scrollingElement || document.documentElement || document.body;
            const describe = (el) => {
                if (!el) return 'window';
                const id = el.id ? `#${el.id}` : '';
                const className = typeof el.className === 'string' && el.className.trim()
                    ? `.${el.className.trim().split(/\\s+/).slice(0, 2).join('.')}`
                    : '';
                return `${(el.tagName || 'window').toLowerCase()}${id}${className}`;
            };
            const seen = new Set();
            const candidates = [];
            const pushCandidate = (el, source) => {
                if (!el || seen.has(el)) return;
                seen.add(el);

                const scrollHeight = Number(el.scrollHeight || 0);
                const clientHeight = Number(el.clientHeight || 0);
                if (scrollHeight <= clientHeight + 48) return;

                const rect = typeof el.getBoundingClientRect === 'function'
                    ? el.getBoundingClientRect()
                    : { height: clientHeight };

                let score = scrollHeight - clientHeight;
                const lastAnchor = anchors.length ? anchors[anchors.length - 1] : null;
                const firstAnchor = anchors.length ? anchors[0] : null;
                if (lastAnchor && typeof el.contains === 'function' && el.contains(lastAnchor)) score += 200000;
                if (firstAnchor && typeof el.contains === 'function' && el.contains(firstAnchor)) score += 50000;
                if (clientHeight >= Math.max((window.innerHeight || clientHeight || 0) * 0.55, 360)) score += 10000;
                if (rect.height >= Math.max((window.innerHeight || rect.height || 0) * 0.45, 280)) score += 5000;
                if (el === fallback || el === document.body || el === document.documentElement) score += 1000;

                candidates.push({ el, score, source });
            };

            pushCandidate(fallback, 'window');
            if (document.body) pushCandidate(document.body, 'body');
            if (document.documentElement) pushCandidate(document.documentElement, 'document');

            const sampleAnchors = anchors.length
                ? [
                    anchors[0],
                    anchors[Math.floor(anchors.length / 2)],
                    anchors[anchors.length - 1],
                ].filter(Boolean)
                : [];

            for (const anchor of sampleAnchors) {
                let node = anchor;
                while (node && node !== document.body) {
                    pushCandidate(node, 'ancestor');
                    node = node.parentElement;
                }
            }

            candidates.sort((a, b) => b.score - a.score);
            const chosen = candidates[0] || { el: fallback, source: 'window' };
            const root = chosen.el || fallback;
            const doc = document.scrollingElement || document.documentElement;
            const unique = new Set(
                anchors.map(anchor => (anchor.href || "").split("?")[0]).filter(Boolean)
            );
            const bodyHeight = Number(Math.max(
                doc.scrollHeight || 0,
                document.body ? document.body.scrollHeight : 0,
                document.documentElement ? document.documentElement.scrollHeight : 0
            ));
            const rootIsWindow = root === fallback || root === document.body || root === document.documentElement;
            const scrollTop = rootIsWindow ? Number(doc.scrollTop || window.pageYOffset || 0) : Number(root.scrollTop || 0);
            const rootHeight = rootIsWindow ? bodyHeight : Number(root.scrollHeight || 0);
            const viewportHeight = rootIsWindow ? Number(window.innerHeight || doc.clientHeight || 0) : Number(root.clientHeight || window.innerHeight || 0);
            return {
                linkCount: unique.size,
                scrollTop,
                scrollHeight: rootHeight,
                rootHeight,
                viewportHeight,
                rootClientHeight: viewportHeight,
                bodyHeight,
                bodyScrollTop: Number(doc.scrollTop || window.pageYOffset || 0),
                scrollRoot: rootIsWindow ? "window" : describe(root),
                scrollRootSource: chosen.source || "window",
                atBottom: scrollTop + viewportHeight >= rootHeight - 24,
            };
        }"""
    )


def advance_profile_grid(page, strategy: str = "anchor-step") -> dict[str, int | bool | str]:
    return page.evaluate(
        """(strategy) => {
            const anchors = Array.from(document.querySelectorAll("a[href*='/p/'], a[href*='/reel/']"));
            const fallback = document.scrollingElement || document.documentElement || document.body;
            const seen = new Set();
            const candidates = [];
            const pushCandidate = (el, source) => {
                if (!el || seen.has(el)) return;
                seen.add(el);
                const scrollHeight = Number(el.scrollHeight || 0);
                const clientHeight = Number(el.clientHeight || 0);
                if (scrollHeight <= clientHeight + 48) return;

                let score = scrollHeight - clientHeight;
                const lastAnchor = anchors.length ? anchors[anchors.length - 1] : null;
                if (lastAnchor && typeof el.contains === 'function' && el.contains(lastAnchor)) score += 200000;
                if (clientHeight >= Math.max((window.innerHeight || clientHeight || 0) * 0.55, 360)) score += 10000;
                if (el === fallback || el === document.body || el === document.documentElement) score += 1000;
                candidates.push({ el, score, source });
            };

            pushCandidate(fallback, 'window');
            if (document.body) pushCandidate(document.body, 'body');
            if (document.documentElement) pushCandidate(document.documentElement, 'document');

            const sampleAnchors = anchors.length
                ? [
                    anchors[0],
                    anchors[Math.floor(anchors.length / 2)],
                    anchors[anchors.length - 1],
                ].filter(Boolean)
                : [];
            for (const anchor of sampleAnchors) {
                let node = anchor;
                while (node && node !== document.body) {
                    pushCandidate(node, 'ancestor');
                    node = node.parentElement;
                }
            }

            candidates.sort((a, b) => b.score - a.score);
            const chosen = candidates[0] || { el: fallback, source: 'window' };
            const root = chosen.el || fallback;
            const doc = document.scrollingElement || document.documentElement;
            const rootIsWindow = root === fallback || root === document.body || root === document.documentElement;
            const beforeTop = rootIsWindow ? Number(doc.scrollTop || window.pageYOffset || 0) : Number(root.scrollTop || 0);
            const viewportHeight = rootIsWindow ? Number(window.innerHeight || doc.clientHeight || 0) : Number(root.clientHeight || window.innerHeight || 0);
            const beforeRootHeight = rootIsWindow
                ? Number(Math.max(doc.scrollHeight || 0, document.body ? document.body.scrollHeight : 0, document.documentElement ? document.documentElement.scrollHeight : 0))
                : Number(root.scrollHeight || 0);
            const beforeBodyHeight = Number(Math.max(
                doc.scrollHeight || 0,
                document.body ? document.body.scrollHeight : 0,
                document.documentElement ? document.documentElement.scrollHeight : 0
            ));

            if (anchors.length) {
                const targetAnchor =
                    strategy === 'anchor-step' ? anchors[anchors.length - 1]
                    : strategy === 'viewport-step' ? anchors[Math.max(0, anchors.length - 4)]
                    : anchors[anchors.length - 1];
                targetAnchor.scrollIntoView({ block: 'end', inline: 'nearest' });
            }

            let step = Math.max(Math.floor((viewportHeight || 900) * 0.9), 900);
            if (strategy === 'viewport-step') {
                step = Math.max(Math.floor((viewportHeight || 900) * 1.4), 1200);
                if (rootIsWindow) {
                    window.scrollBy(0, step);
                } else {
                    root.scrollTop = Number(root.scrollTop || 0) + step;
                    root.dispatchEvent(new Event('scroll', { bubbles: true }));
                }
            } else if (strategy === 'bottom-jump') {
                const target = Math.max(
                    rootIsWindow ? Number(doc.scrollHeight || 0) : Number(root.scrollHeight || 0),
                    document.body ? Number(document.body.scrollHeight || 0) : 0,
                    document.documentElement ? Number(document.documentElement.scrollHeight || 0) : 0
                );
                if (rootIsWindow) {
                    window.scrollTo(0, target);
                } else {
                    root.scrollTop = target;
                    root.dispatchEvent(new Event('scroll', { bubbles: true }));
                }
            } else {
                if (rootIsWindow) {
                    window.scrollBy(0, step);
                } else {
                    root.scrollTop = Number(root.scrollTop || 0) + step;
                    root.dispatchEvent(new Event('scroll', { bubbles: true }));
                }
            }

            const afterScrollTop = rootIsWindow ? Number(doc.scrollTop || window.pageYOffset || 0) : Number(root.scrollTop || 0);
            if (afterScrollTop <= beforeTop + 5) {
                const fallbackTarget = Math.max(
                    Number(doc.scrollHeight || 0),
                    document.body ? Number(document.body.scrollHeight || 0) : 0,
                    document.documentElement ? Number(document.documentElement.scrollHeight || 0) : 0
                );
                if (rootIsWindow) {
                    window.scrollTo(0, fallbackTarget);
                } else {
                    root.scrollTop = Math.max(Number(root.scrollHeight || 0), fallbackTarget);
                    root.dispatchEvent(new Event('scroll', { bubbles: true }));
                }
            }

            window.dispatchEvent(new Event('scroll'));

            const bodyHeight = Number(Math.max(
                Number(doc.scrollHeight || 0),
                document.body ? Number(document.body.scrollHeight || 0) : 0,
                document.documentElement ? Number(document.documentElement.scrollHeight || 0) : 0
            ));
            const rootHeight = rootIsWindow ? bodyHeight : Number(root.scrollHeight || 0);
            return {
                scrollTop: rootIsWindow ? Number(doc.scrollTop || window.pageYOffset || 0) : Number(root.scrollTop || 0),
                scrollHeight: rootHeight,
                rootHeight,
                bodyHeight,
                scrollRoot: rootIsWindow ? "window" : (root.tagName || 'div').toLowerCase(),
                scrollRootSource: chosen.source || "window",
                moved: (rootIsWindow ? Number(doc.scrollTop || window.pageYOffset || 0) : Number(root.scrollTop || 0)) > beforeTop + 24,
                rootHeightGrew: rootHeight > beforeRootHeight + 24,
                bodyHeightGrew: bodyHeight > beforeBodyHeight + 24,
            };
        }""",
        arg=strategy,
    )


def apply_profile_scroll_strategy(page, strategy: str) -> dict[str, int | bool | str]:
    if strategy in {"anchor-step", "viewport-step", "bottom-jump"}:
        return advance_profile_grid(page, strategy)

    if strategy == "mouse-wheel":
        try:
            page.mouse.wheel(0, 2200)
        except Exception:
            pass
        page.wait_for_timeout(120)
        return get_profile_scroll_state(page)

    if strategy == "page-down":
        try:
            page.keyboard.press("PageDown")
        except Exception:
            pass
        page.wait_for_timeout(120)
        return get_profile_scroll_state(page)

    return get_profile_scroll_state(page)


def wait_for_profile_dom_growth(
    page,
    previous_state: dict[str, int | bool | str],
    timeout_ms: int = SCROLL_WAIT_TIMEOUT,
) -> dict[str, int | bool | str]:
    try:
        page.wait_for_function(
            """(prev) => {
                const anchors = Array.from(document.querySelectorAll("a[href*='/p/'], a[href*='/reel/']"));
                const fallback = document.scrollingElement || document.documentElement || document.body;
                const seen = new Set();
                const candidates = [];
                const pushCandidate = (el, source) => {
                    if (!el || seen.has(el)) return;
                    seen.add(el);
                    const scrollHeight = Number(el.scrollHeight || 0);
                    const clientHeight = Number(el.clientHeight || 0);
                    if (scrollHeight <= clientHeight + 48) return;
                    let score = scrollHeight - clientHeight;
                    const lastAnchor = anchors.length ? anchors[anchors.length - 1] : null;
                    if (lastAnchor && typeof el.contains === 'function' && el.contains(lastAnchor)) score += 200000;
                    if (clientHeight >= Math.max((window.innerHeight || clientHeight || 0) * 0.55, 360)) score += 10000;
                    if (el === fallback || el === document.body || el === document.documentElement) score += 1000;
                    candidates.push({ el, score, source });
                };
                pushCandidate(fallback, 'window');
                if (document.body) pushCandidate(document.body, 'body');
                if (document.documentElement) pushCandidate(document.documentElement, 'document');
                const sampleAnchors = anchors.length
                    ? [anchors[0], anchors[Math.floor(anchors.length / 2)], anchors[anchors.length - 1]].filter(Boolean)
                    : [];
                for (const anchor of sampleAnchors) {
                    let node = anchor;
                    while (node && node !== document.body) {
                        pushCandidate(node, 'ancestor');
                        node = node.parentElement;
                    }
                }
                candidates.sort((a, b) => b.score - a.score);
                const root = (candidates[0] || { el: fallback }).el || fallback;
                const doc = document.scrollingElement || document.documentElement;
                const rootIsWindow = root === fallback || root === document.body || root === document.documentElement;
                const unique = new Set(anchors.map(anchor => (anchor.href || "").split("?")[0]).filter(Boolean));
                const scrollTop = rootIsWindow ? Number(doc.scrollTop || window.pageYOffset || 0) : Number(root.scrollTop || 0);
                const bodyHeight = Number(Math.max(
                    doc.scrollHeight || 0,
                    document.body ? document.body.scrollHeight : 0,
                    document.documentElement ? document.documentElement.scrollHeight : 0
                ));
                const scrollHeight = rootIsWindow ? bodyHeight : Number(root.scrollHeight || 0);
                return (
                    unique.size > prev.linkCount ||
                    scrollHeight > prev.scrollHeight + 24 ||
                    scrollHeight > prev.bodyHeight + 24 ||
                    bodyHeight > prev.bodyHeight + 24 ||
                    scrollTop > prev.scrollTop + 24
                );
            }""",
            arg=previous_state,
            timeout=timeout_ms,
        )
    except Exception:
        pass

    return get_profile_scroll_state(page)


def confirm_profile_exhausted(
    page,
    reference_state: dict[str, int | bool | str],
    timeout_ms: int = SCROLL_EXHAUSTION_CONFIRM_TIMEOUT,
) -> tuple[bool, dict[str, int | bool | str]]:
    try:
        page.wait_for_function(
            """(prev) => {
                const anchors = Array.from(document.querySelectorAll("a[href*='/p/'], a[href*='/reel/']"));
                const fallback = document.scrollingElement || document.documentElement || document.body;
                const seen = new Set();
                const candidates = [];
                const pushCandidate = (el) => {
                    if (!el || seen.has(el)) return;
                    seen.add(el);
                    const scrollHeight = Number(el.scrollHeight || 0);
                    const clientHeight = Number(el.clientHeight || 0);
                    if (scrollHeight <= clientHeight + 48) return;
                    let score = scrollHeight - clientHeight;
                    const lastAnchor = anchors.length ? anchors[anchors.length - 1] : null;
                    if (lastAnchor && typeof el.contains === 'function' && el.contains(lastAnchor)) score += 200000;
                    if (clientHeight >= Math.max((window.innerHeight || clientHeight || 0) * 0.55, 360)) score += 10000;
                    if (el === fallback || el === document.body || el === document.documentElement) score += 1000;
                    candidates.push({ el, score });
                };
                pushCandidate(fallback);
                if (document.body) pushCandidate(document.body);
                if (document.documentElement) pushCandidate(document.documentElement);
                const sampleAnchors = anchors.length
                    ? [anchors[0], anchors[Math.floor(anchors.length / 2)], anchors[anchors.length - 1]].filter(Boolean)
                    : [];
                for (const anchor of sampleAnchors) {
                    let node = anchor;
                    while (node && node !== document.body) {
                        pushCandidate(node);
                        node = node.parentElement;
                    }
                }
                candidates.sort((a, b) => b.score - a.score);
                const root = (candidates[0] || { el: fallback }).el || fallback;
                const doc = document.scrollingElement || document.documentElement;
                const rootIsWindow = root === fallback || root === document.body || root === document.documentElement;
                const unique = new Set(
                    anchors.map(anchor => (anchor.href || "").split("?")[0]).filter(Boolean)
                );
                const scrollTop = rootIsWindow ? Number(doc.scrollTop || window.pageYOffset || 0) : Number(root.scrollTop || 0);
                const bodyHeight = Number(Math.max(
                    doc.scrollHeight || 0,
                    document.body ? document.body.scrollHeight : 0,
                    document.documentElement ? document.documentElement.scrollHeight : 0
                ));
                const scrollHeight = rootIsWindow ? bodyHeight : Number(root.scrollHeight || 0);
                return (
                    unique.size > prev.linkCount ||
                    scrollHeight > prev.scrollHeight + 24 ||
                    scrollHeight > prev.bodyHeight + 24 ||
                    bodyHeight > prev.bodyHeight + 24 ||
                    scrollTop > prev.scrollTop + 24
                );
            }""",
            arg=reference_state,
            timeout=timeout_ms,
        )
        return False, get_profile_scroll_state(page)
    except Exception:
        final_state = get_profile_scroll_state(page)
        exhausted = bool(
            final_state["atBottom"]
            and final_state["linkCount"] <= reference_state["linkCount"]
            and final_state["scrollHeight"] <= reference_state["scrollHeight"] + 24
            and final_state["bodyHeight"] <= reference_state["bodyHeight"] + 24
            and final_state["scrollTop"] <= reference_state["scrollTop"] + 24
        )
        return exhausted, final_state


def auto_login_if_needed(page, context, profile_url: str, log_hook: Optional[LogHook] = None) -> bool:
    """Backend-only automatic Instagram login using env vars or isolated test credentials."""
    return login_to_instagram_if_needed(page, context, profile_url, log_hook=log_hook)


def click_button_if_visible(page, pattern: str, timeout_ms: int = 1200) -> bool:
    try:
        button = page.get_by_role("button", name=re.compile(pattern, re.IGNORECASE)).first
        if button.count() > 0:
            button.click(timeout=timeout_ms)
            return True
    except Exception:
        pass

    return False


def save_storage_state(context, log_hook: Optional[LogHook] = None) -> None:
    state_path = get_storage_state_path()
    if state_path is None:
        return

    try:
        state_path.parent.mkdir(parents=True, exist_ok=True)
        context.storage_state(path=str(state_path))
        emit_log(log_hook, "INFO", "Session saved", str(state_path))
    except Exception as exc:
        emit_log(log_hook, "WARN", "Session save skipped", type(exc).__name__)


def login_to_instagram_if_needed(page, context, profile_url: str, log_hook: Optional[LogHook] = None) -> bool:
    if not has_login_credentials():
        return False

    username_selector = "input[name='username']"
    password_selector = "input[name='password']"
    login_form_selector = f"{username_selector}, {password_selector}"

    if not wait_for_selector(page, login_form_selector, LOGIN_FORM_TIMEOUT):
        return False

    emit_log(log_hook, "INFO", "Instagram login", "Login form detected. Attempting secure env-based sign-in.")

    try:
        page.locator(username_selector).first.fill(INSTAGRAM_USERNAME)
        page.locator(password_selector).first.fill(INSTAGRAM_PASSWORD)
        page.locator("button[type='submit']").first.click(timeout=3000)
    except Exception as exc:
        emit_log(log_hook, "WARN", "Instagram login failed", type(exc).__name__)
        return False

    try:
        page.wait_for_load_state("domcontentloaded", timeout=LOGIN_POST_SUBMIT_TIMEOUT)
    except Exception:
        pass

    click_button_if_visible(page, r"not now|skip")
    click_button_if_visible(page, r"save info|save login info")

    page.goto(profile_url, wait_until="domcontentloaded", timeout=POST_GOTO_TIMEOUT)

    if wait_for_selector(page, PROFILE_GRID_SELECTOR, LOGIN_POST_SUBMIT_TIMEOUT):
        save_storage_state(context, log_hook)
        emit_log(log_hook, "SUCCESS", "Instagram login", "Session is ready and profile grid is visible.")
        return True

    emit_log(log_hook, "WARN", "Instagram login incomplete", "Profile grid did not appear after sign-in.")
    return False


def prepare_profile_page(
    page,
    context,
    profile_url: str,
    timeout_ms: int = PROFILE_LINK_WAIT_TIMEOUT,
    log_hook: Optional[LogHook] = None,
) -> None:
    start_time = time.perf_counter()
    page.goto(profile_url, wait_until="domcontentloaded", timeout=POST_GOTO_TIMEOUT)

    initial_wait = min(timeout_ms, PROFILE_LINK_WAIT_TIMEOUT)
    if wait_for_selector(page, PROFILE_GRID_SELECTOR, initial_wait):
        elapsed = time.perf_counter() - start_time
        emit_log(log_hook, "INFO", "Profile ready", f"Grid detected in {elapsed:.2f}s.")
        return

    attempted_login = auto_login_if_needed(page, context, profile_url, log_hook=log_hook)
    if attempted_login:
        elapsed = time.perf_counter() - start_time
        emit_log(log_hook, "INFO", "Profile ready", f"Login-backed session restored in {elapsed:.2f}s.")
        return

    remaining_ms = max(500, timeout_ms - int((time.perf_counter() - start_time) * 1000))
    if wait_for_selector(page, PROFILE_GRID_SELECTOR, remaining_ms):
        elapsed = time.perf_counter() - start_time
        emit_log(log_hook, "INFO", "Profile ready", f"Grid detected after extended wait in {elapsed:.2f}s.")
        return

    raise TimeoutError(
        "Instagram profile grid did not become visible. The profile may require login, "
        "Instagram may be blocking the cloud server, or the page loaded too slowly."
    )


def ask_yes_no(prompt: str) -> bool:
    """Ask a yes/no question until the user gives a clear answer."""
    while True:
        answer = input(prompt).strip().lower()

        if answer in {"y", "yes"}:
            return True
        if answer in {"n", "no"}:
            return False

        print("Please answer with yes or no.")


def normalize_instagram_profile_url(raw_value: str) -> Optional[str]:
    """Validate and normalize an Instagram profile URL for profile-grid scraping."""
    value = raw_value.strip()
    if not value:
        return None

    if not re.match(r"^https?://", value, re.IGNORECASE):
        value = f"https://{value}"

    try:
        parsed = urlparse(value)
    except Exception:
        return None

    hostname = (parsed.hostname or "").lower()
    if parsed.scheme not in {"http", "https"} or hostname not in VALID_INSTAGRAM_HOSTS:
        return None

    path_parts = [part for part in parsed.path.split("/") if part]
    if len(path_parts) != 1:
        return None

    username = path_parts[0]
    if username.lower() in RESERVED_INSTAGRAM_PATHS:
        return None
    if not re.match(r"^[A-Za-z0-9._]{1,30}$", username):
        return None

    return f"https://www.instagram.com/{username}/"


def prompt_instagram_profile_url() -> str:
    """Collect and confirm a valid Instagram profile link."""
    while True:
        raw_value = input("Enter the Instagram profile link: ").strip()
        profile_url = normalize_instagram_profile_url(raw_value)

        if profile_url is None:
            print("Invalid Instagram profile link. Example: https://www.instagram.com/username/")
            continue

        if ask_yes_no(f"You entered: {profile_url}. Do you want to proceed? (yes/no): "):
            return profile_url

        print("Okay, please enter the Instagram link again.")


def prompt_positive_integer(prompt: str) -> int:
    """Collect a positive integer, retrying on blank or invalid input."""
    while True:
        value = input(prompt).strip()

        if not value:
            print("This field cannot be blank. Please enter a positive number.")
            continue
        if not value.isdigit():
            print("Please enter numbers only.")
            continue

        number = int(value)
        if number <= 0:
            print("Please enter a number greater than zero.")
            continue

        return number


def prompt_date(prompt: str) -> datetime:
    """Collect a strict YYYY-MM-DD date."""
    while True:
        value = input(prompt).strip()

        if not value:
            print("Date cannot be blank. Use YYYY-MM-DD, for example 2026-01-01.")
            continue

        try:
            return datetime.strptime(value, DATE_INPUT_FORMAT)
        except ValueError:
            print("Invalid date format. Please use YYYY-MM-DD, for example 2026-01-01.")


def prompt_date_coverage() -> tuple[datetime, Optional[datetime]]:
    """Collect either a start-only coverage or an inclusive start/end date range."""
    while True:
        start_date = prompt_date("Enter start date (YYYY-MM-DD): ")
        end_value = input("Enter end date (YYYY-MM-DD), or type latest: ").strip()

        if not end_value:
            print("End date cannot be blank. Type a date or latest.")
            continue

        if end_value.lower() in {"latest", "l"}:
            return start_date, None

        try:
            end_date = datetime.strptime(end_value, DATE_INPUT_FORMAT)
        except ValueError:
            print("Invalid end date format. Please use YYYY-MM-DD or type latest.")
            continue

        if end_date < start_date:
            print("End date cannot be earlier than the start date. Please enter the date coverage again.")
            continue

        return start_date, end_date


def normalize_excel_filename(raw_value: str) -> Optional[str]:
    """Validate an Excel filename and add .xlsx when the extension is omitted."""
    value = raw_value.strip()
    if not value:
        return None

    if any(char in INVALID_FILENAME_CHARS for char in value):
        return None

    path = Path(value)
    if path.name in {"", ".", ".."}:
        return None
    if path.suffix and path.suffix.lower() != ".xlsx":
        return None

    normalized = path if path.suffix else path.with_suffix(".xlsx")
    if not normalized.name or normalized.stem.strip(" .") == "":
        return None

    return str(normalized)


def prompt_excel_filename() -> str:
    """Collect a valid output workbook filename and confirm overwrite when needed."""
    while True:
        raw_value = input("Enter Excel output filename: ").strip()
        filename = normalize_excel_filename(raw_value)

        if filename is None:
            print("Invalid filename. Use a normal Excel name like instagram_report.xlsx.")
            print("Do not use these characters: < > : \" / \\ | ? *")
            continue

        if Path(filename).exists():
            if ask_yes_no(f"{filename} already exists. Overwrite it? (yes/no): "):
                return filename

            print("Okay, please enter a new Excel filename.")
            continue

        return filename


def format_date_coverage(start_date: datetime, end_date: Optional[datetime]) -> str:
    start_text = start_date.strftime(DATE_INPUT_FORMAT)
    if end_date is None:
        return f"{start_text} to latest post"

    return f"{start_text} to {end_date.strftime(DATE_INPUT_FORMAT)}"


def show_config_summary(config: ScrapeConfig) -> None:
    print("\nScraping setup summary")
    print("-" * 60)
    print(f"Instagram link: {config.profile_url}")
    print(f"Scroll rounds:   {config.scroll_rounds}")
    print(f"Date coverage:   {format_date_coverage(config.start_date, config.end_date)}")
    print(f"Excel file:      {config.output_file}")
    print("-" * 60)


def prompt_scrape_config() -> ScrapeConfig:
    """Collect all required settings and ask for final confirmation before scraping."""
    while True:
        profile_url = prompt_instagram_profile_url()
        scroll_rounds = prompt_positive_integer("Enter number of scroll rounds: ")
        start_date, end_date = prompt_date_coverage()
        output_file = prompt_excel_filename()

        config = ScrapeConfig(
            profile_url=profile_url,
            scroll_rounds=scroll_rounds,
            start_date=start_date,
            end_date=end_date,
            output_file=output_file,
        )

        show_config_summary(config)
        if ask_yes_no("Proceed with scraping? (yes/no): "):
            return config

        print("No problem. Let's update the scraping inputs.\n")


def post_matches_date_coverage(post: PostData, start_date: datetime, end_date: Optional[datetime]) -> bool:
    """Only include posts whose detected date is within the requested range."""
    return classify_post_date_coverage(post.post_date_obj, start_date, end_date)[0] == "in_range"


def classify_post_date_coverage(
    post_date_obj: Optional[datetime],
    start_date: datetime,
    end_date: Optional[datetime],
) -> tuple[str, str]:
    if post_date_obj is None:
        return "unknown_date", "Skipped because the post date could not be detected reliably."

    post_date = post_date_obj.date()
    if post_date < start_date.date():
        return "older_than_start", f"Skipped because {post_date.strftime(DATE_INPUT_FORMAT)} is older than start date {start_date.strftime(DATE_INPUT_FORMAT)}."
    if end_date is not None and post_date > end_date.date():
        return "newer_than_end", f"Skipped because {post_date.strftime(DATE_INPUT_FORMAT)} is newer than end date {end_date.strftime(DATE_INPUT_FORMAT)}."

    if end_date is None:
        return "in_range", f"Included because {post_date.strftime(DATE_INPUT_FORMAT)} is on/after start date {start_date.strftime(DATE_INPUT_FORMAT)}."

    return "in_range", (
        f"Included because {post_date.strftime(DATE_INPUT_FORMAT)} is within "
        f"{start_date.strftime(DATE_INPUT_FORMAT)} to {end_date.strftime(DATE_INPUT_FORMAT)}."
    )


def parse_count(value: Optional[str]) -> Optional[int]:
    if not value:
        return None

    text = value.strip().upper().replace(",", "")
    match = re.match(r"^(\d+(?:\.\d+)?)([KMB]?)$", text)
    if not match:
        digits = re.sub(r"[^\d]", "", text)
        return int(digits) if digits else None

    number = float(match.group(1))
    suffix = match.group(2)

    if suffix == "K":
        number *= 1000
    elif suffix == "M":
        number *= 1000000
    elif suffix == "B":
        number *= 1000000000

    return int(number)




def extract_counts_from_json_ld(content: Optional[str]) -> tuple[Optional[int], Optional[int]]:
    if not content:
        return None, None

    try:
        data = json.loads(content)
    except Exception:
        return None, None

    likes: Optional[int] = None
    comments: Optional[int] = None

    def walk(node) -> None:
        nonlocal likes, comments

        if isinstance(node, list):
            for item in node:
                walk(item)
            return

        if not isinstance(node, dict):
            return

        if comments is None and "commentCount" in node:
            comments = parse_count(str(node.get("commentCount")))

        interaction_stats = node.get("interactionStatistic")
        if isinstance(interaction_stats, dict):
            interaction_stats = [interaction_stats]

        if isinstance(interaction_stats, list):
            for stat in interaction_stats:
                if not isinstance(stat, dict):
                    continue

                interaction_type = str(stat.get("interactionType", ""))
                if "LikeAction" in interaction_type and likes is None:
                    likes = parse_count(str(stat.get("userInteractionCount")))

        for value in node.values():
            if likes is not None and comments is not None:
                return
            walk(value)

    walk(data)
    return likes, comments


def extract_post_payload_from_next_data(page) -> Optional[dict[str, Any]]:
    try:
        raw = page.locator("script#__NEXT_DATA__").first.text_content(timeout=NEXT_DATA_TIMEOUT)
        if not raw:
            return None

        data = json.loads(raw)
    except Exception:
        return None

    best: Optional[dict[str, Any]] = None

    def walk(node: Any) -> None:
        nonlocal best
        if best is not None:
            return

        if isinstance(node, list):
            for item in node:
                walk(item)
            return

        if not isinstance(node, dict):
            return

        has_like_block = isinstance(node.get("edge_media_preview_like"), dict) or isinstance(node.get("edge_liked_by"), dict)
        has_comment_block = isinstance(node.get("edge_media_to_parent_comment"), dict) or isinstance(node.get("edge_media_to_comment"), dict)
        has_type = "__typename" in node or "is_video" in node

        # Prefer the first node that clearly looks like Instagram post media payload.
        if has_type and (has_like_block or has_comment_block):
            best = node
            return

        for value in node.values():
            walk(value)

    walk(data)
    return best


def infer_post_type(url: str, payload: Optional[dict[str, Any]]) -> str:
    lowered = url.lower()
    if "/reel/" in lowered:
        return "Reel"
    if "/tv/" in lowered:
        return "Video"

    typename = str(payload.get("__typename", "")) if payload else ""
    if "Sidecar" in typename:
        return "Carousel"
    if "Video" in typename:
        return "Video"
    if "Image" in typename:
        return "Photo"

    if "/p/" in lowered:
        return "Photo/Video"

    return "Unknown"


def infer_post_type_from_dom(page, url: str) -> str:
    lowered = url.lower()
    if "/reel/" in lowered:
        return "Reel"
    if "/tv/" in lowered:
        return "Video"

    try:
        if page.locator("article video").count() > 0:
            return "Video"
    except Exception:
        pass

    try:
        if page.locator("article svg[aria-label='Next'], article button[aria-label='Next']").count() > 0:
            return "Carousel"
    except Exception:
        pass

    return "Photo/Video"


def extract_counts_from_next_data_payload(payload: Optional[dict[str, Any]]) -> tuple[Optional[int], Optional[int], Optional[int]]:
    if not payload:
        return None, None, None

    likes: Optional[int] = None
    comments: Optional[int] = None
    shares: Optional[int] = None

    # Try multiple keys for likes (Instagram changed structure multiple times)
    for like_key in ["edge_media_preview_like", "edge_liked_by", "likeCount", "likes"]:
        like_block = payload.get(like_key)
        if isinstance(like_block, dict) and likes is None:
            likes = parse_count(str(like_block.get("count")))
            if likes is not None:
                break
    
    if likes is None and isinstance(payload.get("likeCount"), int):
        likes = payload.get("likeCount")

    # Try multiple keys for comments
    for comment_key in ["edge_media_to_parent_comment", "edge_media_to_comment", "commentCount", "comments"]:
        comment_block = payload.get(comment_key)
        if isinstance(comment_block, dict) and comments is None:
            comments = parse_count(str(comment_block.get("count")))
            if comments is not None:
                break
    
    if comments is None and isinstance(payload.get("commentCount"), int):
        comments = payload.get("commentCount")

    # Try multiple keys for shares (often not exposed, so None is acceptable)
    for share_key in ["share_count", "reshare_count", "shareCount", "reshareCount", "shares"]:
        if share_key in payload and shares is None:
            shares = parse_count(str(payload.get(share_key)))
            if shares is not None:
                break

    return likes, comments, shares


def extract_counts_from_text(content: Optional[str]) -> tuple[Optional[int], Optional[int], Optional[int]]:
    if not content:
        return None, None, None

    normalized = " ".join(content.split())

    likes: Optional[int] = None
    comments: Optional[int] = None
    shares: Optional[int] = None

    # Likes: Match various formats like "123 likes", "1.2K likes", "1,234 likes"
    like_match = re.search(r"([\d.,KMkm]+)\s+likes?\b", normalized, re.IGNORECASE)
    if like_match:
        likes = parse_count(like_match.group(1))

    # Comments: "View all 123 comments" or "456 comments"
    comment_patterns = [
        r"View all\s+([\d.,KMkm]+)\s+comments?\b",
        r"([\d.,KMkm]+)\s+comments?\b",
    ]
    for pattern in comment_patterns:
        for comment_match in re.finditer(pattern, normalized, re.IGNORECASE):
            context = normalized[comment_match.start():comment_match.end() + 40]
            if re.search(r"comments?\s+from\s+facebook\b", context, re.IGNORECASE):
                continue

            comments = parse_count(comment_match.group(1))
            break

        if comments is not None:
            break

    # Shares: "123 shares" - Note: shares are rarely exposed publicly
    share_patterns = [
        r"([\d.,KMkm]+)\s+shares?\b",
        r"([\d.,KMkm]+)\s+shared\b",
    ]
    for pattern in share_patterns:
        share_match = re.search(pattern, normalized, re.IGNORECASE)
        if share_match:
            shares = parse_count(share_match.group(1))
            break

    return likes, comments, shares


def extract_metric_text_candidates(page, include_body: bool = False) -> List[str]:
    candidates: List[str] = []
    seen: set[str] = set()

    def add_candidate(value: Optional[str]) -> None:
        if not value:
            return

        normalized = " ".join(value.split())
        if not normalized or normalized in seen:
            return

        seen.add(normalized)
        candidates.append(normalized)

    # Prefer metric-adjacent scopes for both speed and correctness.
    for selector in ["article [role='menuitem']", "article ul[role='menubar']", "article section", "article"]:
        try:
            locator = page.locator(selector).first
            if locator.count() > 0:
                add_candidate(locator.inner_text(timeout=SCOPED_TEXT_TIMEOUT))
        except Exception:
            pass

    for selector in ["meta[property='og:description']", "meta[name='description']"]:
        try:
            add_candidate(page.locator(selector).get_attribute("content"))
        except Exception:
            pass

    try:
        scripts = page.locator("script[type='application/ld+json']").evaluate_all(
            "els => els.map(el => el.textContent || '')"
        )
        for script in scripts:
            add_candidate(script)
    except Exception:
        pass

    tail_selectors = ["body"] if include_body else []
    for selector in tail_selectors:
        try:
            locator = page.locator(selector).first
            if locator.count() > 0:
                add_candidate(locator.inner_text(timeout=BODY_TEXT_TIMEOUT))
        except Exception:
            pass

    return candidates


VISIBLE_ACTION_METRICS_SCRIPT = r"""
() => {
    const normalize = (value) => (value || "").replace(/\u00a0/g, " ").trim();
    const metricNumberPattern = /^\d[\d.,]*(?:\s*[KMB])?$/i;

    const isVisible = (el) => {
        if (!el || !el.isConnected) return false;
        const style = window.getComputedStyle(el);
        if (style.display === "none" || style.visibility === "hidden" || style.opacity === "0") {
            return false;
        }
        const rect = el.getBoundingClientRect();
        return rect.width > 0 && rect.height > 0;
    };

    const rectInfo = (el) => {
        const rect = el.getBoundingClientRect();
        return {
            left: rect.left,
            right: rect.right,
            top: rect.top,
            bottom: rect.bottom,
            width: rect.width,
            height: rect.height,
            centerX: rect.left + rect.width / 2,
            centerY: rect.top + rect.height / 2,
        };
    };

    const metricFromLabel = (label) => {
        const text = normalize(label).toLowerCase();
        if (!text) return null;
        if (/\b(unlike|like)\b/.test(text)) return "like";
        if (/\bcomment\b/.test(text)) return "comment";
        if (/\brepost\b/.test(text)) return "repost";
        if (/\bshare\b/.test(text)) return "share";
        return null;
    };

    const nearestActionElement = (icon) => {
        return icon.closest("[role='button'], button, a") || icon;
    };

    const collectActions = (row) => {
        const byMetric = {};
        const icons = Array.from(row.querySelectorAll("svg[aria-label], [aria-label]"));

        for (const icon of icons) {
            if (!isVisible(icon)) continue;
            const metric = metricFromLabel(icon.getAttribute("aria-label") || icon.textContent);
            if (!metric || byMetric[metric]) continue;

            const action = nearestActionElement(icon);
            if (!isVisible(action)) continue;
            byMetric[metric] = {
                metric,
                label: normalize(icon.getAttribute("aria-label") || icon.textContent),
                rect: rectInfo(action),
            };
        }

        return byMetric;
    };

    const collectCounts = (row) => {
        const counts = [];
        const elements = Array.from(row.querySelectorAll("span, div, a, button"));

        for (const el of elements) {
            if (!isVisible(el)) continue;

            const text = normalize(el.innerText || el.textContent);
            if (!metricNumberPattern.test(text)) continue;

            const rect = rectInfo(el);
            if (rect.width > 140 || rect.height > 60) continue;

            counts.push({ text: text.replace(/\\s+/g, ""), rect });
        }

        return counts;
    };

    const rowCandidates = Array.from(document.querySelectorAll("main section, article section, article, section"))
        .filter(isVisible)
        .map((row) => {
            const actions = collectActions(row);
            const metrics = Object.keys(actions);
            const counts = collectCounts(row);
            const rowRect = rectInfo(row);
            const hasRequired = metrics.includes("like") && metrics.includes("comment");
            const score =
                (hasRequired ? 0 : 1000) +
                Math.abs(rowRect.height - 44) +
                Math.max(0, rowRect.height - 120) * 5 +
                Math.max(0, counts.length - 4) * 20 +
                Math.max(0, metrics.length - 4) * 10;

            return { row, actions, counts, score, metrics };
        })
        .filter((candidate) => candidate.metrics.includes("like") || candidate.metrics.includes("comment"))
        .sort((a, b) => a.score - b.score);

    const selected = rowCandidates[0];
    if (!selected) return {};

    const orderedActions = Object.values(selected.actions).sort((a, b) => a.rect.centerX - b.rect.centerX);
    const usedCounts = new Set();
    const result = {};

    for (const action of orderedActions) {
        if (!["like", "comment", "repost"].includes(action.metric)) continue;

        const nextAction = orderedActions.find((candidate) => candidate.rect.centerX > action.rect.centerX + 1);
        const maxRight = nextAction ? nextAction.rect.left : action.rect.right + 120;

        const possibleCounts = selected.counts
            .map((count, index) => ({ ...count, index }))
            .filter((count) => !usedCounts.has(count.index))
            .filter((count) => count.rect.centerX >= action.rect.right - 4)
            .filter((count) => count.rect.centerX <= maxRight + 8)
            .filter((count) => Math.abs(count.rect.centerY - action.rect.centerY) <= 24)
            .map((count) => ({
                ...count,
                score:
                    Math.max(0, count.rect.left - action.rect.right) +
                    Math.abs(count.rect.centerY - action.rect.centerY) * 4,
            }))
            .sort((a, b) => a.score - b.score);

        if (possibleCounts.length > 0) {
            result[action.metric] = possibleCounts[0].text;
            usedCounts.add(possibleCounts[0].index);
        }
    }

    return result;
}
"""


def extract_visible_metrics(page) -> tuple[Optional[int], Optional[int], Optional[int], int]:
    """Read visible metrics by pairing each action icon with the nearest count in the same row."""
    try:
        result = page.evaluate(VISIBLE_ACTION_METRICS_SCRIPT)
    except Exception:
        result = {}

    if not isinstance(result, dict):
        result = {}

    likes = parse_count(str(result.get("like"))) if result.get("like") else None
    comments = parse_count(str(result.get("comment"))) if result.get("comment") else None
    reposts = parse_count(str(result.get("repost"))) if result.get("repost") else None

    # Instagram often shows a Share icon without a public share count.
    shares = reposts if reposts is not None else 0
    return likes, comments, reposts, shares


def extract_post_metrics(page) -> tuple[Optional[int], Optional[int], Optional[int], Optional[dict[str, Any]]]:
    """Extract likes, comments, shares using 3-layer priority:
    1. Visible metrics (source of truth - what user sees)
    2. Structured payload (__NEXT_DATA__)
    3. Text extraction (og:description, JSON-LD, page text)
    """
    likes: Optional[int] = None
    comments: Optional[int] = None
    shares: Optional[int] = None
    payload: Optional[dict[str, Any]] = None
    
    extraction_log = []

    # ========== LAYER 1: VISIBLE METRICS (Source of Truth) ==========
    visible_likes, visible_comments, visible_reposts, visible_shares = extract_visible_metrics(page)
    if visible_likes is not None:
        likes = visible_likes
        extraction_log.append(f"visible_likes={likes}")
    if visible_comments is not None:
        comments = visible_comments
        extraction_log.append(f"visible_comments={comments}")
    if visible_reposts is not None:
        extraction_log.append(f"visible_reposts={visible_reposts}")
    if visible_shares is not None and visible_shares > 0:
        shares = visible_shares
        extraction_log.append(f"visible_shares={shares}")

    # Short-circuit when primary source already resolved the required fields.
    if likes is not None and comments is not None:
        if extraction_log:
            print(f"      Extracted: {', '.join(extraction_log)}")
        return likes, comments, shares, payload
    
    # ========== LAYER 2: STRUCTURED PAYLOAD ==========
    payload = extract_post_payload_from_next_data(page)
    payload_likes, payload_comments, payload_shares = extract_counts_from_next_data_payload(payload)
    
    if likes is None and payload_likes is not None:
        likes = payload_likes
        extraction_log.append(f"payload_likes={likes}")
    if comments is None and payload_comments is not None:
        comments = payload_comments
        extraction_log.append(f"payload_comments={comments}")
    if payload_shares is not None:
        shares = payload_shares
        extraction_log.append(f"payload_shares={shares}")

    # Short-circuit when payload completed remaining fields.
    if likes is not None and comments is not None:
        if extraction_log:
            print(f"      Extracted: {', '.join(extraction_log)}")
        return likes, comments, shares, payload

    # ========== LAYER 3: TEXT EXTRACTION (Fallback) ==========
    if likes is None or comments is None:
        scoped_candidates = extract_metric_text_candidates(page, include_body=False)

        for content in scoped_candidates:
            text_likes, text_comments, text_shares = extract_counts_from_text(content)

            if likes is None and text_likes is not None:
                likes = text_likes
                extraction_log.append(f"text_likes={likes}")
            if comments is None and text_comments is not None:
                comments = text_comments
                extraction_log.append(f"text_comments={comments}")
            if shares is None and text_shares is not None:
                shares = text_shares
                extraction_log.append(f"text_shares={shares}")

            # Stop if we found likes and comments
            if likes is not None and comments is not None:
                break

            # Try JSON-LD extraction as well
            json_likes, json_comments = extract_counts_from_json_ld(content)
            if likes is None and json_likes is not None:
                likes = json_likes
                extraction_log.append(f"json_likes={likes}")
            if comments is None and json_comments is not None:
                comments = json_comments
                extraction_log.append(f"json_comments={comments}")

        # Last-resort broad scan only if still unresolved.
        if likes is None or comments is None:
            body_candidates = extract_metric_text_candidates(page, include_body=True)
            broad_candidates = [content for content in body_candidates if content not in scoped_candidates]

            for content in broad_candidates:
                text_likes, text_comments, text_shares = extract_counts_from_text(content)
                if likes is None and text_likes is not None:
                    likes = text_likes
                    extraction_log.append(f"broad_text_likes={likes}")
                if comments is None and text_comments is not None:
                    comments = text_comments
                    extraction_log.append(f"broad_text_comments={comments}")
                if shares is None and text_shares is not None:
                    shares = text_shares
                    extraction_log.append(f"broad_text_shares={shares}")
                if likes is not None and comments is not None:
                    break

    # Log extraction sources for debugging
    if extraction_log:
        print(f"      Extracted: {', '.join(extraction_log)}")
    
    return likes, comments, shares, payload


def parse_datetime(date_str: str) -> Optional[datetime]:
    if not date_str:
        return None

    formats = [
        "%Y-%m-%dT%H:%M:%S.%fZ",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%d",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue

    return None


def wait_for_profile_ready(page) -> None:
    """Ensures profile page has loaded with post grid."""
    if wait_for_selector(page, PROFILE_GRID_SELECTOR, PROFILE_LINK_WAIT_TIMEOUT):
        return

    wait_for_selector(page, PROFILE_GRID_SELECTOR, PROFILE_RETRY_MS)


def wait_for_post_ready(page, url: str) -> None:
    """Ensures post page has loaded with metadata (date, metrics)."""
    try:
        page.wait_for_load_state("domcontentloaded", timeout=1200)
    except Exception:
        pass

    try:
        page.locator("time").first.wait_for(timeout=POST_TIME_WAIT_TIMEOUT)
    except Exception:
        pass

    try:
        page.locator("article").first.wait_for(timeout=POST_ARTICLE_WAIT_TIMEOUT)
    except Exception:
        pass


def wait_for_metric_elements(page) -> bool:
    """Wait for metric elements (likes/comments/shares) to appear on page."""
    if wait_for_selector(page, METRIC_SELECTOR, METRIC_READY_TIMEOUT):
        return True

    return wait_for_selector(
        page,
        "meta[property='og:description'], script#__NEXT_DATA__, script[type='application/ld+json']",
        METRIC_FALLBACK_MS,
    )


def extract_date(page) -> tuple[str, Optional[datetime]]:
    try:
        time_el = page.locator("time").first
        dt = time_el.get_attribute("datetime")
        if dt:
            return dt, parse_datetime(dt)

        visible = time_el.inner_text().strip()
        return visible, parse_datetime(visible)
    except Exception:
        return "", None


def open_post_for_extraction(page, url: str, goto_timeout: int = POST_GOTO_TIMEOUT) -> tuple[str, Optional[datetime], str]:
    page.goto(url, wait_until="domcontentloaded", timeout=goto_timeout)
    wait_for_post_ready(page, url)
    post_type = infer_post_type_from_dom(page, url)
    raw_date, date_obj = extract_date(page)
    return raw_date, date_obj, post_type


def probe_post_date(
    page,
    url: str,
    log_hook: Optional[LogHook] = None,
    goto_timeout: int = 15000,
) -> Optional[datetime]:
    try:
        _, date_obj, _ = open_post_for_extraction(page, url, goto_timeout=goto_timeout)
        return date_obj
    except Exception as exc:
        emit_log(log_hook, "WARN", "Oldest date probe failed", f"{url} ({type(exc).__name__})")
        return None


def extract_metrics_from_loaded_post(
    page,
    url: str,
    raw_date: str,
    date_obj: Optional[datetime],
    post_type: str,
    log_hook: Optional[LogHook] = None,
) -> PostData:
    likes: Optional[int] = None
    comments: Optional[int] = None
    shares = 0

    wait_for_metric_elements(page)
    extracted_likes, extracted_comments, extracted_shares, payload = extract_post_metrics(page)
    if payload is not None and post_type == "Photo/Video":
        post_type = infer_post_type(url, payload)

    likes = extracted_likes
    comments = extracted_comments
    shares = extracted_shares if extracted_shares is not None else 0

    emit_log(
        log_hook,
        "INFO",
        "Metrics extracted",
        f"{url} -> likes={likes}, comments={comments}, shares={shares}, date={raw_date or 'N/A'}",
    )

    return PostData(
        url=url,
        post_type=post_type,
        post_date_raw=raw_date,
        post_date_obj=date_obj,
        likes=likes,
        comments=comments,
        shares=shares,
    )


def collect_post_links(
    page,
    max_posts: Optional[int] = None,
    scroll_rounds: int = MAX_SCROLL_ROUNDS,
    target_start_date: Optional[datetime] = None,
    probe_page=None,
    log_hook: Optional[LogHook] = None,
    progress_hook: Optional[ProgressHook] = None,
    live_hook: Optional[LiveHook] = None,
    cancel_check: Optional[Callable[[], bool]] = None,
    diagnostics: Optional[dict[str, Any]] = None,
) -> List[str]:
    links = {}  # Use dict instead of set to preserve insertion order (Python 3.7+)
    stagnant = 0
    no_more_posts_proof = 0
    stopped_early = False
    coverage_reached = False
    likely_exhausted_logged = False
    stagnant_limit = get_effective_stagnant_limit(scroll_rounds)
    min_rounds_before_stop = get_minimum_rounds_before_stop(scroll_rounds, stagnant_limit)
    proof_rounds_required = SCROLL_STOP_PROOF_ROUNDS
    coverage_mode = target_start_date is not None
    oldest_visible_date_seen: Optional[datetime] = None
    newest_visible_date_seen: Optional[datetime] = None
    last_probed_oldest_link = ""

    if coverage_mode:
        stagnant_limit = max(stagnant_limit, 12)
        min_rounds_before_stop = scroll_rounds
        proof_rounds_required = max(proof_rounds_required, 3)

    stop_reason = f"Reached max scroll rounds ({scroll_rounds})."

    wait_for_profile_ready(page)
    print(
        "Starting collection: "
        f"max_rounds={scroll_rounds}, stagnant_limit={stagnant_limit}, "
        f"proof_rounds={proof_rounds_required}, min_rounds_before_stop={min_rounds_before_stop}"
    )
    emit_log(
        log_hook,
        "INFO",
        "Collection config",
        (
            f"max_rounds={scroll_rounds}, stagnant_limit={stagnant_limit}, "
            f"proof_rounds={proof_rounds_required}, coverage_start="
            f"{target_start_date.strftime(DATE_INPUT_FORMAT) if target_start_date else 'none'}."
        ),
    )
    if coverage_mode:
        emit_log(
            log_hook,
            "INFO",
            "Continuing scroll to reach target date",
            f"Using conservative early-stop thresholds to reach posts as old as {target_start_date.strftime(DATE_INPUT_FORMAT)}.",
        )

    initial_found = collect_visible_post_links(page)
    for href in initial_found:
        if href and href not in links:
            links[href] = True

    if links:
        print(f"  Initial grid: +{len(links)} links (total={len(links)})")
        emit_log(log_hook, "INFO", "Initial grid", f"+{len(links)} new links (total={len(links)}).")
        if probe_page is not None:
            newest_link = next(iter(links))
            newest_date = probe_post_date(probe_page, newest_link, log_hook=log_hook)
            if newest_date is not None:
                newest_visible_date_seen = newest_date
                emit_log(log_hook, "INFO", "Newest post so far", newest_date.strftime(DATE_INPUT_FORMAT))
            initial_oldest_link = next(reversed(links))
            oldest_date = probe_post_date(probe_page, initial_oldest_link, log_hook=log_hook)
            if oldest_date is not None:
                oldest_visible_date_seen = oldest_date
                if newest_visible_date_seen is None:
                        newest_visible_date_seen = oldest_date
                last_probed_oldest_link = initial_oldest_link
                emit_log(log_hook, "INFO", "Oldest post so far", oldest_date.strftime(DATE_INPUT_FORMAT))
    if live_hook is not None:
        try:
            live_hook(
                page,
                "initial-grid",
                {
                    "round": 0,
                    "newLinks": len(links),
                    "totalLinks": len(links),
                    "oldestVisibleDate": oldest_visible_date_seen.strftime(DATE_INPUT_FORMAT) if oldest_visible_date_seen else "",
                    "newestVisibleDate": newest_visible_date_seen.strftime(DATE_INPUT_FORMAT) if newest_visible_date_seen else "",
                },
            )
        except Exception:
            pass
    emit_progress(progress_hook, 0, scroll_rounds, len(links))

    for scroll_round in range(1, scroll_rounds + 1):
        if cancel_check is not None and cancel_check():
            raise RuntimeError("Cancelled during profile scrolling.")

        round_started = time.perf_counter()
        round_start_state = get_profile_scroll_state(page)
        previous_state = round_start_state
        current_state = round_start_state
        before = len(links)
        dom_changed = False
        scrolled = False
        height_grew = False
        movement_grew = False
        at_bottom_with_no_growth = False
        exhaustion_confirmed = False
        strategy_used = "none"
        attempts_used = 0

        for attempt_index in range(SCROLL_ATTEMPTS_PER_ROUND):
            strategy_used = SCROLL_STRATEGIES[min(attempt_index, len(SCROLL_STRATEGIES) - 1)]
            attempts_used = attempt_index + 1

            try:
                apply_profile_scroll_strategy(page, strategy_used)
                current_state = wait_for_profile_dom_growth(page, previous_state, SCROLL_WAIT_TIMEOUT)
                found = collect_visible_post_links(page)
            except Exception:
                print(f"  Scroll {scroll_round}: Failed to inspect grid using {strategy_used}")
                emit_log(
                    log_hook,
                    "WARN",
                    f"Scroll {scroll_round}",
                    f"Failed to inspect profile grid after {strategy_used}.",
                )
                page.wait_for_timeout(PROFILE_RETRY_MS)
                continue

            for href in found:
                if href:
                    clean_url = href.split("?")[0]
                    if clean_url not in links:
                        links[clean_url] = True  # Add to dict to preserve order

            scrolled = scrolled or bool(current_state["scrollTop"] > previous_state["scrollTop"] + 24)
            height_grew = height_grew or bool(current_state["scrollHeight"] > previous_state["scrollHeight"] + 24)
            movement_grew = movement_grew or bool(current_state["linkCount"] > previous_state["linkCount"])
            dom_changed = dom_changed or scrolled or height_grew or movement_grew

            if len(links) > before:
                break

            at_bottom_with_no_growth = bool(
                previous_state["atBottom"]
                and current_state["atBottom"]
                and current_state["scrollHeight"] <= previous_state["scrollHeight"] + 24
                and current_state["scrollTop"] <= previous_state["scrollTop"] + 24
            )
            if at_bottom_with_no_growth:
                exhaustion_confirmed, current_state = confirm_profile_exhausted(page, current_state)
            previous_state = current_state

        new_count = len(links) - before
        scroll_details = (
            f"strategy={strategy_used}, attempts={attempts_used}, root={current_state['scrollRoot']}({current_state['scrollRootSource']}), "
            f"root_height={round_start_state['scrollHeight']}->{current_state['scrollHeight']}, "
            f"body_height={round_start_state['bodyHeight']}->{current_state['bodyHeight']}, "
            f"top={round_start_state['scrollTop']}->{current_state['scrollTop']}, "
            f"visible_anchors={round_start_state['linkCount']}->{current_state['linkCount']}, "
            f"total_links={before}->{len(links)}"
        )

        if new_count > 0:
            stagnant = 0
            no_more_posts_proof = 0
            print(f"  Scroll {scroll_round}: +{new_count} new links (total={len(links)})")
            emit_log(
                log_hook,
                "INFO",
                f"Scroll {scroll_round}",
                f"+{new_count} new links (total={len(links)}, {scroll_details}).",
            )
            if probe_page is not None:
                oldest_link = next(reversed(links))
                if oldest_link != last_probed_oldest_link:
                    oldest_date = probe_post_date(probe_page, oldest_link, log_hook=log_hook)
                    if oldest_date is not None:
                        last_probed_oldest_link = oldest_link
                        oldest_visible_date_seen = oldest_date
                        newest_visible_date_seen = newest_visible_date_seen or oldest_date
                        emit_log(log_hook, "INFO", "Oldest post so far", oldest_date.strftime(DATE_INPUT_FORMAT))
                        if target_start_date is not None and oldest_date.date() <= target_start_date.date():
                            coverage_reached = True
                            emit_log(
                                log_hook,
                                "SUCCESS",
                                "Reached target date while scrolling",
                                (
                                    f"Oldest visible post {oldest_date.strftime(DATE_INPUT_FORMAT)} is on/before "
                                    f"start date {target_start_date.strftime(DATE_INPUT_FORMAT)}."
                                ),
                            )
        else:
            if dom_changed:
                stagnant = 0
                no_more_posts_proof = 0
                print(f"  Scroll {scroll_round}: +0 new links but content moved (total={len(links)})")
                emit_log(
                    log_hook,
                    "INFO",
                    f"Scroll {scroll_round}",
                    f"+0 new links; content moved ({scroll_details}, moved={scrolled}, height_grew={height_grew or current_state.get('rootHeightGrew')}, body_height_grew={current_state.get('bodyHeightGrew')}).",
                )
            else:
                stagnant += 1
                if exhaustion_confirmed:
                    no_more_posts_proof += 1
                    reason = (
                        f"+0 new links; at bottom with no DOM growth after confirmation wait "
                        f"(proof {no_more_posts_proof}/{proof_rounds_required}, "
                        f"stagnant {stagnant}/{stagnant_limit}, total={len(links)})."
                    )
                    print(f"  Scroll {scroll_round}: {reason}")
                elif at_bottom_with_no_growth:
                    no_more_posts_proof = 0
                    reason = (
                        f"+0 new links; bottom reached but late lazy loading appeared possible "
                        f"(continuing, stagnant {stagnant}/{stagnant_limit}, total={len(links)})."
                    )
                    print(f"  Scroll {scroll_round}: {reason}")
                elif not scrolled:
                    no_more_posts_proof = 0
                    reason = (
                        f"+0 new links; scroll did not move after {strategy_used} "
                        f"({scroll_details}, lazy loading may have stalled, stagnant {stagnant}/{stagnant_limit}, total={len(links)})."
                    )
                    print(f"  Scroll {scroll_round}: {reason}")
                else:
                    no_more_posts_proof = 0
                    reason = (
                        f"+0 new links after {strategy_used} "
                        f"({scroll_details}, stagnant {stagnant}/{stagnant_limit}, total={len(links)})."
                    )
                    print(f"  Scroll {scroll_round}: {reason}")

                emit_log(
                    log_hook,
                    "INFO",
                    f"Scroll {scroll_round}",
                    reason,
                )

        emit_progress(progress_hook, scroll_round, scroll_rounds, len(links))
        if live_hook is not None:
            try:
                live_hook(
                    page,
                    "scroll-round",
                    {
                        "round": scroll_round,
                        "newLinks": new_count,
                        "totalLinks": len(links),
                        "stagnant": stagnant,
                        "oldestVisibleDate": oldest_visible_date_seen.strftime(DATE_INPUT_FORMAT) if oldest_visible_date_seen else "",
                        "newestVisibleDate": newest_visible_date_seen.strftime(DATE_INPUT_FORMAT) if newest_visible_date_seen else "",
                        "strategy": strategy_used,
                        "reason": reason if new_count == 0 else "",
                    },
                )
            except Exception:
                pass

        if coverage_reached:
            stop_reason = (
                f"Reached target start date {target_start_date.strftime(DATE_INPUT_FORMAT)} while scrolling; "
                f"oldest visible post is {oldest_visible_date_seen.strftime(DATE_INPUT_FORMAT) if oldest_visible_date_seen else 'unknown'}."
            )
            emit_log(log_hook, "INFO", "Link collection stopped", stop_reason)
            stopped_early = True
            break

        if max_posts is not None and len(links) >= max_posts:
            print(f"Reached max_posts limit: {max_posts}")
            emit_log(log_hook, "INFO", "Link collection stopped", f"Reached max_posts limit: {max_posts}.")
            stop_reason = f"Reached max_posts limit: {max_posts}."
            stopped_early = True
            break

        if coverage_mode:
            if (
                scroll_round >= max(3, scroll_rounds // 2)
                and stagnant >= stagnant_limit
                and no_more_posts_proof >= proof_rounds_required
                and not likely_exhausted_logged
            ):
                emit_log(
                    log_hook,
                    "WARN",
                    "Likely exhausted before target date",
                    (
                        f"Scroll stalled with strong exhaustion proof, but the collector will continue until "
                        f"max rounds ({scroll_rounds}) because target date "
                        f"{target_start_date.strftime(DATE_INPUT_FORMAT) if target_start_date else 'n/a'} has not been confirmed yet."
                    ),
                )
                likely_exhausted_logged = True
        elif (
            scroll_round >= min_rounds_before_stop
            and stagnant >= stagnant_limit
            and no_more_posts_proof >= proof_rounds_required
        ):
            print(
                f"Stopping: strong proof of exhaustion after {scroll_round} rounds "
                f"(stagnant={stagnant}/{stagnant_limit}, proof={no_more_posts_proof}/{proof_rounds_required})"
            )
            emit_log(
                log_hook,
                "INFO",
                "Link collection stopped",
                f"Reached strong end-of-profile proof after {scroll_round} rounds (stagnant={stagnant}, proof={no_more_posts_proof}).",
            )
            stop_reason = (
                f"Reached strong end-of-profile proof after {scroll_round} rounds "
                f"(stagnant={stagnant}, proof={no_more_posts_proof})."
            )
            stopped_early = True
            break

        elapsed = time.perf_counter() - round_started
        if elapsed >= SLOW_SCROLL_SECONDS:
            print(f"  Scroll {scroll_round}: Slow round ({elapsed:.2f}s)")
            emit_log(log_hook, "WARN", "Slow scroll", f"Round {scroll_round} took {elapsed:.2f}s.")

    if not stopped_early and scroll_rounds > 0:
        emit_log(
            log_hook,
            "INFO",
            "Link collection stopped",
            f"Reached max scroll rounds ({scroll_rounds}) with {len(links)} unique links collected.",
        )

    if live_hook is not None:
        try:
            live_hook(
                page,
                "scroll-stop",
                {
                    "round": min(scroll_rounds, scroll_round if scroll_rounds > 0 else 0),
                    "totalLinks": len(links),
                    "reason": stop_reason,
                    "coverageReached": coverage_reached,
                    "oldestVisibleDate": oldest_visible_date_seen.strftime(DATE_INPUT_FORMAT) if oldest_visible_date_seen else "",
                    "newestVisibleDate": newest_visible_date_seen.strftime(DATE_INPUT_FORMAT) if newest_visible_date_seen else "",
                },
            )
        except Exception:
            pass

    if diagnostics is not None:
        diagnostics.clear()
        diagnostics.update(
            {
                "totalLinksCollected": len(links),
                "oldestVisibleDate": oldest_visible_date_seen,
                "newestVisibleDate": newest_visible_date_seen,
                "stopReason": stop_reason,
                "coverageReached": coverage_reached,
                "stagnantCount": stagnant,
                "exhaustionProof": no_more_posts_proof,
            }
        )

    print(f"Collection complete: {len(links)} unique links found\n")
    return list(links.keys())[:max_posts]  # Convert dict keys to list in insertion order


def extract_post_data(page, url: str, log_hook: Optional[LogHook] = None) -> PostData:
    post_type = "Unknown"
    raw_date = ""
    date_obj: Optional[datetime] = None
    likes: Optional[int] = None
    comments: Optional[int] = None
    shares: int = 0  # Default to 0 if not found

    for attempt in range(1, POST_LOAD_RETRIES + 1):
        attempt_started = time.perf_counter()
        try:
            raw_date, date_obj, post_type = open_post_for_extraction(page, url)
            post_data = extract_metrics_from_loaded_post(page, url, raw_date, date_obj, post_type, log_hook=log_hook)
            likes = post_data.likes
            comments = post_data.comments
            shares = post_data.shares
            post_type = post_data.post_type

            elapsed = time.perf_counter() - attempt_started
            
            # Log what we found
            found_metrics = []
            if likes is not None:
                found_metrics.append(f"likes={likes}")
            if comments is not None:
                found_metrics.append(f"comments={comments}")
            found_metrics.append(f"shares={shares}")
            
            if found_metrics:
                print(f"    Attempt {attempt}: Extracted {', '.join(found_metrics)} in {elapsed:.2f}s")
            else:
                print(f"    Attempt {attempt}: No metrics extracted after {elapsed:.2f}s")

            if elapsed >= SLOW_POST_SECONDS:
                emit_log(log_hook, "WARN", "Slow post", f"{url} took {elapsed:.2f}s on attempt {attempt}.")

            missing_critical_metrics = likes is None and comments is None
            if missing_critical_metrics and attempt < POST_LOAD_RETRIES:
                emit_log(log_hook, "WARN", "Retrying post", f"{url} missing likes/comments on attempt {attempt}.")
                wait_time = 600 if attempt == 1 else 1500
                page.wait_for_timeout(wait_time)
                continue

            emit_log(
                log_hook,
                "INFO",
                f"Attempt {attempt}",
                f"{url} -> likes={likes}, comments={comments}, shares={shares}, date={raw_date or 'N/A'} ({elapsed:.2f}s)",
            )
            break
            
        except Exception as e:
            elapsed = time.perf_counter() - attempt_started
            print(f"    Attempt {attempt}: Error - {type(e).__name__}")
            emit_log(log_hook, "WARN", f"Attempt {attempt} failed", f"{url} ({type(e).__name__}, {elapsed:.2f}s)")
            if attempt == POST_LOAD_RETRIES:
                # Final attempt failed - mark with "Cannot detect"
                post_type = "Unknown"
                raw_date = "Cannot detect"
                date_obj = None
                likes = None
                comments = None
                shares = 0  # Default to 0
            else:
                # Wait before retrying, with exponential backoff
                wait_time = 600 if attempt == 1 else 1500
                print(f"    Retrying in {wait_time}ms...")
                page.wait_for_timeout(wait_time)

    return PostData(
        url=url,
        post_type=post_type,
        post_date_raw=raw_date,
        post_date_obj=date_obj,
        likes=likes,
        comments=comments,
        shares=shares,
    )


def format_count(value: Optional[int]) -> str:
    return "Cannot detect" if value is None else str(value)


def format_month_total(values: List[Optional[int]]) -> str:
    if any(value is None for value in values):
        return "Cannot detect"

    return str(sum(value for value in values if value is not None))


def format_post_date(post: PostData) -> str:
    if post.post_date_obj:
        return post.post_date_obj.strftime("%m/%d/%Y")

    return post.post_date_raw or "Cannot detect"


def save_grouped_excel(posts: List[PostData], filename: str, coverage_label: str = "starting Jan - latest"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Instagram Data 2026"

    # top title row
    ws["A1"] = coverage_label
    ws["A2"] = "Blank values mean Instagram did not expose that exact count on the public post page."

    # headers
    ws["A4"] = "Link to Post"
    ws["B4"] = "Date Posted"
    ws["C4"] = "Post Type"
    ws["D4"] = "No. of Comments"
    ws["E4"] = "No. of Likes"
    ws["F4"] = "No. of Share"

    grouped = defaultdict(list)
    for post in posts:
        if post.post_date_obj:
            month_name = post.post_date_obj.strftime("%B")
            grouped[month_name].append(post)
        else:
            grouped["Unable to detect date"].append(post)

    month_order = ["January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December",
                   "Unable to detect date"]

    current_row = 5

    for month in month_order:
        month_posts = grouped.get(month, [])
        if not month_posts:
            continue

        # month title row
        ws[f"A{current_row}"] = month

        ws[f"D{current_row}"] = format_month_total([p.comments for p in month_posts])
        ws[f"E{current_row}"] = format_month_total([p.likes for p in month_posts])
        ws[f"F{current_row}"] = format_month_total([p.shares for p in month_posts])

        current_row += 1

        # sort posts by date
        month_posts.sort(key=lambda p: p.post_date_obj or datetime.min)

        for post in month_posts:
            ws[f"A{current_row}"] = post.url
            ws[f"B{current_row}"] = format_post_date(post)
            ws[f"C{current_row}"] = post.post_type
            ws[f"D{current_row}"] = format_count(post.comments)
            ws[f"E{current_row}"] = format_count(post.likes)
            ws[f"F{current_row}"] = format_count(post.shares)
            current_row += 1

        current_row += 2  # blank rows before next month

    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    wb.save(filename)


def save_empty_result_excel(
    filename: str,
    coverage_label: str,
    total_links_collected: int,
    oldest_detected: Optional[datetime],
    newest_detected: Optional[datetime],
    reason: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Instagram Data 2026"

    ws["A1"] = coverage_label
    ws["A2"] = "No posts found within selected date range."
    ws["A4"] = "Selected date coverage"
    ws["B4"] = coverage_label
    ws["A5"] = "Total links collected"
    ws["B5"] = str(total_links_collected)
    ws["A6"] = "Newest detected post"
    ws["B6"] = newest_detected.strftime(DATE_INPUT_FORMAT) if newest_detected else "Unknown"
    ws["A7"] = "Oldest detected post"
    ws["B7"] = oldest_detected.strftime(DATE_INPUT_FORMAT) if oldest_detected else "Unknown"
    ws["A8"] = "Reason"
    ws["B8"] = reason

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    wb.save(filename)


def route_nonessential_resources(route) -> None:
    if route.request.resource_type in BLOCKED_RESOURCE_TYPES:
        route.abort()
        return

    route.continue_()


def install_playwright_browsers() -> None:
    """Install Chromium browser binaries when a cloud build skipped them."""
    command = [
        sys.executable,
        "-m",
        "playwright",
        "install",
        "chromium",
        "chromium-headless-shell",
    ]
    print("Installing missing Playwright browser binaries...")
    subprocess.run(command, check=True, timeout=240)


def launch_browser(playwright):
    """Create a cloud-safe Playwright browser/context pair."""
    launch_options = {
        "headless": PLAYWRIGHT_HEADLESS,
        "args": [
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-dev-shm-usage",
            "--disable-gpu",
            "--no-first-run",
            "--no-default-browser-check",
        ],
    }

    try:
        browser = playwright.chromium.launch(**launch_options)
    except PlaywrightError as exc:
        message = str(exc)
        missing_browser = "Executable doesn't exist" in message or "Please run the following command" in message
        if not (PLAYWRIGHT_AUTO_INSTALL and missing_browser):
            raise

        install_playwright_browsers()
        browser = playwright.chromium.launch(**launch_options)

    context_options = {
        "viewport": {"width": 1400, "height": 900},
        "locale": "en-US",
        "timezone_id": "Asia/Manila",
        "ignore_https_errors": True,
    }

    state_path = get_storage_state_path(require_exists=True)
    if state_path is not None:
        context_options["storage_state"] = str(state_path)

    context = browser.new_context(**context_options)
    return browser, context


def run_scrape(context, config: ScrapeConfig) -> None:
    """Run the existing scraping flow with an already-created browser context."""
    context.route("**/*", route_nonessential_resources)

    page = context.new_page()

    if PLAYWRIGHT_HEADLESS:
        print("Running in headless mode. Manual Instagram login prompt is skipped.")
        if PLAYWRIGHT_STORAGE_STATE:
            print(f"Using Playwright storage state: {PLAYWRIGHT_STORAGE_STATE}")
    else:
        page.goto(config.profile_url, wait_until="domcontentloaded", timeout=POST_GOTO_TIMEOUT)
        input("Log in to Instagram if needed, then press Enter... ")

    prepare_profile_page(page, context, config.profile_url)

    links = collect_post_links(
        page,
        MAX_POSTS,
        config.scroll_rounds,
        target_start_date=config.start_date,
    )
    print(f"Found {len(links)} posts.")

    all_posts = []
    old_posts_count = 0
    missing_metrics_count = 0
    post_delay = BASE_POST_DELAY
    consecutive_misses = 0

    for i, link in enumerate(links, start=1):
        print(f"[{i}/{len(links)}] Processing: {link}")
        post = extract_post_data(page, link)
        all_posts.append(post)

        missing = []
        if post.likes is None:
            missing.append("likes")
        if post.comments is None:
            missing.append("comments")

        if missing:
            missing_metrics_count += 1
            consecutive_misses += 1
            post_delay = min(MAX_POST_DELAY, BASE_POST_DELAY + 0.1 * consecutive_misses)
            print(f"    Slow load: missing {', '.join(missing)} - will show Cannot detect")
        else:
            consecutive_misses = 0
            post_delay = BASE_POST_DELAY
            print("    Metrics loaded")

        time.sleep(post_delay)

    if missing_metrics_count > 0:
        pct = round(100 * missing_metrics_count / len(links), 1)
        print(f"\n{missing_metrics_count}/{len(links)} posts ({pct}%) had slow-loading metrics.")
        print("These posts are still included with Cannot detect for missing metrics.\n")

    filtered_posts = []
    for post in all_posts:
        if post_matches_date_coverage(post, config.start_date, config.end_date):
            filtered_posts.append(post)
        else:
            old_posts_count += 1

    if old_posts_count > 0:
        coverage = format_date_coverage(config.start_date, config.end_date)
        print(f"Filtered out {old_posts_count} posts outside {coverage}.")

    coverage = format_date_coverage(config.start_date, config.end_date)
    print(f"Processing {len(filtered_posts)} posts within valid date range.")
    save_grouped_excel(filtered_posts, config.output_file, coverage)
    print(f"Saved to {config.output_file}")


def main():
    config = prompt_scrape_config()

    with sync_playwright() as p:
        browser, context = launch_browser(p)
        try:
            run_scrape(context, config)
        finally:
            context.close()
            browser.close()

if __name__ == "__main__":
    main()
